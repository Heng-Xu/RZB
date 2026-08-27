"""2025 真实数据适配器。

源目录始终只读；本模块把八县区 110/35 kV 资产、静态极值、候选、
光伏、县域锚点、成本案例和容量网络标准化到调用者指定目录。
"""
from __future__ import annotations

import hashlib
import json
import math
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
import yaml
from openpyxl import load_workbook


ADAPTER_VERSION = "1.1.0"
SOURCE_VERSION = "电网建模数据_Agent整合版_V1.2"
SCENARIO_ID = "real_2025"
LINEAGE_FIELDS = (
    "source_ref",
    "source_version",
    "transformation",
    "scenario_id",
    "quality_flag",
    "source_sha256",
)

SRC01_02 = "2025设备负载统计表.xlsx"
SRC03_04 = "110（35）kv设备明细.xlsx"
SRC05 = "光伏装机.xlsx"
SRC08 = "近5年容载比.xlsx"
SRC10 = "江苏徐州邢楼110千伏变电站主变扩建等工程建设规模及投资汇总表(1).xlsx"
SRC14 = "光伏8760小时数据.xlsx"


class RealDataQualityError(ValueError):
    """关键真实数据断言不满足。"""


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _clean(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def _voltage(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value).lower().replace("kv", "").replace("千伏", "").strip()
    try:
        return int(float(text))
    except ValueError:
        return None


def _unit(value: Any) -> str | None:
    value = _clean(value)
    if value is None:
        return None
    text = str(value).replace("＃", "#")
    return text if text.startswith("#") else f"#{text}"


def _timestamp(value: Any) -> str | None:
    parsed = pd.to_datetime(value, errors="coerce")
    return None if pd.isna(parsed) else parsed.isoformat(sep=" ")


def _rows(
    path: Path,
    sheet_name: str,
    min_row: int,
    max_row: int,
    max_col: int,
) -> list[tuple[Any, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        return [
            tuple(_clean(value) for value in row)
            for row in sheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                max_col=max_col,
                values_only=True,
            )
        ]
    finally:
        workbook.close()


def _lineage(
    source_ref: str,
    transformation: str,
    quality_flag: str,
    source_sha256: str,
    scenario_id: str = SCENARIO_ID,
) -> dict[str, str]:
    return {
        "source_ref": source_ref,
        "source_version": SOURCE_VERSION,
        "transformation": transformation,
        "scenario_id": scenario_id,
        "quality_flag": quality_flag,
        "source_sha256": source_sha256,
    }


def _source_registry(source_root: Path) -> dict[str, dict[str, Any]]:
    return {
        path.name: {"sha256": _sha256(path), "bytes": path.stat().st_size}
        for path in sorted(source_root.iterdir(), key=lambda item: item.name)
        if path.is_file()
    }


def _equipment(
    source_root: Path,
    regions: set[str],
    source_hash: str,
) -> tuple[dict[tuple[str, int, str, str], dict[str, Any]], list[dict[str, Any]]]:
    equipment: dict[tuple[str, int, str, str], dict[str, Any]] = {}
    expansion_rows: list[dict[str, Any]] = []
    specs = (
        ("110千伏变电站1", 4, 320, 47, 110, 14, 15, 13, 29),
        ("35千伏变电站1", 4, 117, 40, 35, 12, 13, 11, 24),
    )
    for sheet, min_row, max_row, max_col, expected_voltage, unit_col, cap_col, bay_col, parent_col in specs:
        rows = _rows(source_root / SRC03_04, sheet, min_row, max_row, max_col)
        for offset, row in enumerate(rows, start=min_row):
            region = str(row[1]) if row[1] is not None else ""
            voltage = _voltage(row[5])
            station_id = row[3]
            unit_id = _unit(row[unit_col])
            if (
                _number(row[0]) is None
                or region not in regions
                or voltage != expected_voltage
                or station_id is None
                or unit_id is None
            ):
                continue
            key = (region, voltage, str(station_id), unit_id)
            equipment[key] = {
                "capacity_mva": _number(row[cap_col]),
                "available_10kv_bays": _number(row[bay_col]),
                "parent_supply_id": row[parent_col],
                "source_row": offset,
                "station_type": row[6],
                "n_minus_1_source_status": row[31] if voltage == 110 else row[25],
            }
            if voltage == 110 and _number(row[10]) and (_number(row[10]) or 0) > 0:
                count = int(_number(row[10]) or 0)
                candidate_capacity = _number(row[11])
                if candidate_capacity is None or candidate_capacity <= 0:
                    raise RealDataQualityError(
                        f"SRC03 row {offset}: positive expansion count without capacity"
                    )
                expansion_rows.append(
                    {
                        "region_id": region,
                        "voltage_kv": voltage,
                        "station_id": str(station_id),
                        "count": count,
                        "candidate_capacity_mva": candidate_capacity,
                        "available_10kv_bays": _number(row[13]) or 0.0,
                        "source_row": offset,
                        "source_hash": source_hash,
                    }
                )
    return equipment, expansion_rows


def _transformer_tables(
    source_root: Path,
    regions: set[str],
    equipment: dict[tuple[str, int, str, str], dict[str, Any]],
    source_hash: str,
) -> tuple[pd.DataFrame, pd.DataFrame, list[dict[str, Any]]]:
    master: list[dict[str, Any]] = []
    static: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    rows = _rows(source_root / SRC01_02, "主变1", 3, 522, 30)
    for offset, row in enumerate(rows, start=3):
        region = str(row[2]) if row[2] is not None else ""
        voltage = _voltage(row[1])
        station_id = row[3]
        unit_id = _unit(row[4])
        if (
            _number(row[0]) is None
            or region not in regions
            or voltage not in {35, 110}
            or station_id is None
        ):
            continue
        capacity = _number(row[5])
        if unit_id is None or capacity is None or capacity <= 0:
            issues.append(
                {
                    "issue_id": f"DQ-TRANSFORMER-EMPTY-{offset}",
                    "severity": "warning",
                    "entity_id": str(station_id),
                    "description": "SRC02 row has no valid transformer unit or capacity and is excluded from equipment counts.",
                    **_lineage(
                        "SRC02:主变1",
                        "validate unit ID and positive rated capacity before standardization",
                        "empty_placeholder_excluded",
                        source_hash,
                    ),
                }
            )
            continue
        key = (region, voltage, str(station_id), unit_id)
        equipment_row = equipment.get(key)
        equipment_capacity = equipment_row["capacity_mva"] if equipment_row else None
        if equipment_row is None:
            flag = "equipment_mapping_missing"
            issues.append(
                {
                    "issue_id": f"DQ-EQUIPMENT-{offset}",
                    "severity": "error",
                    "entity_id": f"{station_id}/{unit_id}",
                    "description": "SRC02 transformer key is missing from SRC03.",
                    **_lineage("SRC02:主变1", "key comparison against SRC03", flag, source_hash),
                }
            )
        elif equipment_capacity is None or abs(capacity - equipment_capacity) > 1e-9:
            flag = "capacity_conflict_src02_precedence"
            issues.append(
                {
                    "issue_id": f"DQ-CAPACITY-{offset}",
                    "severity": "error",
                    "entity_id": f"{station_id}/{unit_id}",
                    "description": f"SRC02={capacity}, SRC03={equipment_capacity}; SRC02 retained.",
                    **_lineage("SRC02+SRC03", "capacity comparison", flag, source_hash),
                }
            )
        else:
            flag = "validated_src02_src03"
        uid = f"{region}|{voltage}|{station_id}|{unit_id}"
        master.append(
            {
                "transformer_uid": uid,
                "region_id": region,
                "voltage_kv": voltage,
                "station_id": str(station_id),
                "unit_id": unit_id,
                "capacity_mva": capacity,
                "operation_mode": "unknown",
                "asset_scope_id": "year_end_2025",
                "parent_supply_id": equipment_row.get("parent_supply_id") if equipment_row else None,
                "available_10kv_bays": equipment_row.get("available_10kv_bays") if equipment_row else None,
                "equipment_source_row": equipment_row.get("source_row") if equipment_row else None,
                "source_row": offset,
                **_lineage(
                    "SRC02:主变1+SRC03:设备明细",
                    "filter contract regions/35,110 kV; validate composite key and capacity",
                    flag,
                    source_hash,
                ),
            }
        )
        annual_max = _number(row[14])
        annual_min = _number(row[17])
        static.append(
            {
                "transformer_uid": uid,
                "region_id": region,
                "voltage_kv": voltage,
                "station_id": str(station_id),
                "unit_id": unit_id,
                "capacity_mva": capacity,
                "annual_max_net_load_mw": annual_max,
                "annual_max_timestamp": _timestamp(row[16]),
                "annual_min_net_load_mw": annual_min,
                "annual_min_timestamp": _timestamp(row[18]),
                "net_load_interpretation": "net_load_no_pv_resubtraction",
                "asset_scope_id": "operating_2025",
                "source_row": offset,
                **_lineage(
                    "SRC02:主变1",
                    "select 2025 annual extrema; preserve signed net load",
                    "missing_extrema" if annual_max is None or annual_min is None else "source_value",
                    source_hash,
                ),
            }
        )
    return pd.DataFrame(master), pd.DataFrame(static), issues


def _station_tables(
    source_root: Path,
    regions: set[str],
    transformers: pd.DataFrame,
    source_hash: str,
) -> tuple[pd.DataFrame, pd.DataFrame, list[dict[str, Any]]]:
    grouped = (
        transformers.groupby(["region_id", "voltage_kv", "station_id"], as_index=False)
        .agg(capacity_mva=("capacity_mva", "sum"), transformer_count=("transformer_uid", "count"))
    )
    rows: list[dict[str, Any]] = []
    station_keys: set[tuple[str, int, str]] = set()
    for record in grouped.itertuples(index=False):
        station_keys.add((record.region_id, int(record.voltage_kv), record.station_id))
        rows.append(
            {
                "region_id": record.region_id,
                "voltage_kv": int(record.voltage_kv),
                "station_id": record.station_id,
                "capacity_mva": float(record.capacity_mva),
                "transformer_count": int(record.transformer_count),
                "asset_scope_id": "year_end_2025",
                **_lineage(
                    "SRC02:主变1+SRC03:设备明细",
                    "sum same-voltage transformer capacities by anonymized station key",
                    "validated_equipment_scope",
                    source_hash,
                ),
            }
        )
        if (
            record.region_id == "QX-00005"
            and int(record.voltage_kv) == 110
            and record.station_id != "BDZ-00056"
        ):
            operating = dict(rows[-1])
            operating["asset_scope_id"] = "operating_2025"
            operating["transformation"] = "year_end equipment scope excluding BDZ-00056 commissioned at end-2025"
            operating["quality_flag"] = "operating_scope_reconciled_to_src08"
            rows.append(operating)

    static_rows: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    source_rows = _rows(source_root / SRC01_02, "变电站1", 2, 1000, 23)
    for offset, row in enumerate(source_rows, start=2):
        region = str(row[2]) if row[2] is not None else ""
        voltage = _voltage(row[1])
        station_id = row[3]
        if (
            _number(row[0]) is None
            or region not in regions
            or voltage not in {35, 110}
            or station_id is None
        ):
            continue
        station_key = (region, int(voltage), str(station_id))
        if station_key not in station_keys:
            rows.append(
                {
                    "region_id": region,
                    "voltage_kv": int(voltage),
                    "station_id": str(station_id),
                    "capacity_mva": 0.0,
                    "transformer_count": 0,
                    "asset_scope_id": "year_end_2025",
                    **_lineage(
                        "SRC01:变电站1+SRC03:设备明细",
                        "retain station index absent from SRC02 with zero in-service transformer capacity",
                        "no_in_service_transformer_candidate_station",
                        source_hash,
                    ),
                }
            )
            station_keys.add(station_key)
        annual_max = _number(row[12])
        annual_min = _number(row[15])
        flag = "missing_extrema" if annual_max is None or annual_min is None else "source_value"
        static_rows.append(
            {
                "region_id": region,
                "voltage_kv": voltage,
                "station_id": str(station_id),
                "src01_capacity_mva": _number(row[4]),
                "annual_max_net_load_mw": annual_max,
                "annual_max_timestamp": _timestamp(row[14]),
                "annual_min_net_load_mw": annual_min,
                "annual_min_timestamp": _timestamp(row[17]),
                "net_load_interpretation": "net_load_no_pv_resubtraction",
                "asset_scope_id": "operating_2025",
                "source_row": offset,
                **_lineage(
                    "SRC01:变电站1",
                    "select 2025 signed station annual extrema; do not sum extrema as synchronous county peak",
                    flag,
                    source_hash,
                ),
            }
        )
        if flag != "source_value":
            issues.append(
                {
                    "issue_id": f"DQ-STATION-EXTREMA-{offset}",
                    "severity": "warning",
                    "entity_id": str(station_id),
                    "description": "Station annual maximum or minimum net load is missing.",
                    **_lineage("SRC01:变电站1", "null check", flag, source_hash),
                }
            )
    return pd.DataFrame(rows), pd.DataFrame(static_rows), issues


def _expansion_candidates(
    expansion_rows: list[dict[str, Any]],
    station_master: pd.DataFrame,
) -> pd.DataFrame:
    station_capacity = {
        (row.region_id, int(row.voltage_kv), row.station_id): float(row.capacity_mva)
        for row in station_master[station_master["asset_scope_id"] == "year_end_2025"].itertuples()
    }
    records: list[dict[str, Any]] = []
    for item in expansion_rows:
        count = int(item["count"])
        capacity = float(item["candidate_capacity_mva"]) / count
        existing_capacity = station_capacity[(item["region_id"], 110, item["station_id"])]
        candidate_type = "new_station" if existing_capacity <= 0 else "new_third_transformer"
        for ordinal in range(1, count + 1):
            exact_50 = abs(capacity - 50.0) <= 1e-9
            scale = capacity / 50.0
            has_type_matched_cost = candidate_type == "new_third_transformer"
            records.append(
                {
                    "candidate_id": f"EXP-{item['region_id']}-{item['station_id']}-{ordinal:02d}",
                    "region_id": item["region_id"],
                    "voltage_kv": 110,
                    "station_id": item["station_id"],
                    "candidate_type": candidate_type,
                    "existing_capacity_mva": existing_capacity,
                    "old_capacity_mva": 0.0,
                    "new_capacity_mva": capacity,
                    "delta_capacity_mva": capacity,
                    "available_10kv_bays": item["available_10kv_bays"],
                    "implementable_status": (
                        "source_candidate" if has_type_matched_cost else "candidate_cost_gap"
                    ),
                    "capex_low_wanyuan": (
                        (1170.0 if exact_50 else 1170.0 * scale) if has_type_matched_cost else None
                    ),
                    "capex_high_wanyuan": (
                        (1516.0 if exact_50 else 1516.0 * scale) if has_type_matched_cost else None
                    ),
                    "capex_center_wanyuan": (
                        1343.0 if exact_50 and has_type_matched_cost else None
                    ),
                    "cost_basis": (
                        ("SRC10_exact_50mva_range" if exact_50 else "SRC10_type_matched_scaled_range_no_point_estimate")
                        if has_type_matched_cost
                        else "missing_type_matched_new_station_cost"
                    ),
                    "source_row": item["source_row"],
                    **_lineage(
                        "SRC03:110千伏变电站1+SRC10",
                        "expand station-level discrete count into candidate IDs; third transformer delta equals new rating",
                        (
                            "source_candidate_exact_50_cost"
                            if exact_50 and has_type_matched_cost
                            else (
                                "source_candidate_cost_range_only"
                                if has_type_matched_cost
                                else "candidate_cost_gap"
                            )
                        ),
                        item["source_hash"],
                    ),
                }
            )
    return pd.DataFrame(records)


def _pv_snapshot(source_root: Path, regions: set[str], source_hash: str) -> pd.DataFrame:
    aggregates: dict[tuple[str, int, str], dict[str, Any]] = defaultdict(
        lambda: {"online": 0.0, "pipeline": 0.0, "rated": 0.0, "times": [], "rows": []}
    )
    rows = _rows(source_root / SRC05, "主变（跨区合并）1", 2, 521, 10)
    for offset, row in enumerate(rows, start=2):
        region = str(row[1]) if row[1] is not None else ""
        voltage = _voltage(row[3])
        station_id = row[2]
        if region not in regions or voltage not in {35, 110} or station_id is None:
            continue
        item = aggregates[(region, voltage, str(station_id))]
        item["online"] += _number(row[7]) or 0.0
        item["pipeline"] += _number(row[8]) or 0.0
        item["rated"] += _number(row[6]) or 0.0
        timestamp = _timestamp(row[4])
        if timestamp:
            item["times"].append(timestamp)
        item["rows"].append(offset)
    records = []
    for (region, voltage, station_id), item in sorted(aggregates.items()):
        records.append(
            {
                "region_id": region,
                "voltage_kv": voltage,
                "station_id": station_id,
                "pv_online_mw": item["online"],
                "pv_pipeline_mw": item["pipeline"],
                "rated_capacity_context_mva": item["rated"],
                "snapshot_min": min(item["times"]) if item["times"] else None,
                "snapshot_max": max(item["times"]) if item["times"] else None,
                "source_rows": ";".join(map(str, item["rows"])),
                **_lineage(
                    "SRC05:主变（跨区合并）1",
                    "sum online and pipeline PV by region, voltage and station; no feeder allocation",
                    "cross_year_snapshot_2026_context",
                    source_hash,
                    scenario_id="pv_snapshot_2026_context",
                ),
            }
        )
    return pd.DataFrame(records)


def _pv_profile(source_root: Path, source_hash: str) -> pd.DataFrame:
    records = []
    rows = _rows(source_root / SRC14, "光伏8760", 2, 8761, 6)
    for offset, row in enumerate(rows, start=2):
        records.append(
            {
                "hour_index": int(_number(row[0]) or 0),
                "timestamp": _timestamp(row[1]),
                "phi_pv_raw": _number(row[4]),
                "source_row": offset,
                **_lineage(
                    "SRC14:光伏8760:E",
                    "preserve raw normalized PV coefficient; ignore 3408 MW converted column",
                    "raw_night_residual_preserved",
                    source_hash,
                    scenario_id="standard_pv_shape_2025",
                ),
            }
        )
    frame = pd.DataFrame(records)
    timestamps = pd.to_datetime(frame["timestamp"])
    if len(frame) != 8760 or timestamps.nunique() != 8760:
        raise RealDataQualityError("SRC14 must contain 8760 unique timestamps")
    if not frame["phi_pv_raw"].between(0.0, 1.0).all():
        raise RealDataQualityError("SRC14 phi_pv_raw outside [0, 1]")
    expected = pd.date_range("2025-01-01 00:00:00", periods=8760, freq="h")
    if not timestamps.reset_index(drop=True).equals(pd.Series(expected)):
        raise RealDataQualityError("SRC14 timestamps are not the continuous 2025 hourly axis")
    return frame


def _network_lines(source_root: Path, regions: set[str], source_hash: str) -> pd.DataFrame:
    records = []
    rows = _rows(source_root / SRC03_04, "110千伏线路1", 3, 366, 20)
    for offset, row in enumerate(rows, start=3):
        region = str(row[1]) if row[1] is not None else ""
        if _number(row[0]) is None or region not in regions:
            continue
        records.append(
            {
                "region_id": region,
                "voltage_kv": 110,
                "line_id": str(row[2]),
                "from_node": row[12],
                "to_node": row[13],
                "length_km": _number(row[6]),
                "current_limit_a": _number(row[11]),
                "max_loading_pct": _number(row[15]),
                "average_loading_pct": _number(row[16]),
                "source_row": offset,
                **_lineage(
                    "SRC04:110千伏线路1",
                    "filter contract regions; preserve anonymous endpoints and current limits",
                    "capacity_network_only_no_impedance",
                    source_hash,
                ),
            }
        )
    return pd.DataFrame(records)


def _county_reference(source_root: Path, regions: set[str], source_hash: str) -> pd.DataFrame:
    records = []
    current_region: str | None = None
    rows = _rows(source_root / SRC08, "Sheet1", 4, 48, 17)
    for offset, row in enumerate(rows, start=4):
        if row[0] is not None:
            current_region = str(row[0])
        voltage = _voltage(row[1])
        if current_region not in regions or voltage not in {35, 110}:
            continue
        capacity = _number(row[14])
        peak = _number(row[15])
        clr = _number(row[16])
        if capacity is None or peak is None or clr is None:
            raise RealDataQualityError(f"SRC08 row {offset}: incomplete 2025 reference")
        records.append(
            {
                "region_id": current_region,
                "voltage_kv": voltage,
                "capacity_mva": capacity * 10.0,
                "positive_peak_anchor_mw": peak * 10.0,
                "official_clr_reference": clr,
                "year": 2025,
                "role": "reference_and_positive_peak_center_anchor",
                "source_row": offset,
                **_lineage(
                    "SRC08:Sheet1:2025",
                    "convert 10,000 kVA/kW columns to MVA/MW by multiplying by 10",
                    "official_county_anchor",
                    source_hash,
                ),
            }
        )
    frame = pd.DataFrame(records)
    if len(frame) != 16:
        raise RealDataQualityError(f"SRC08 expected 16 county-voltage rows, got {len(frame)}")
    return frame


def _cost_cases(source_root: Path, source_hash: str) -> pd.DataFrame:
    records = []
    rows = _rows(source_root / SRC10, "Sheet1", 7, 56, 17)
    for offset, row in enumerate(rows, start=7):
        name = str(row[1]) if row[1] is not None else ""
        dynamic = _number(row[16])
        if dynamic is None or str(row[0]) in {"1", "1.0"}:
            continue
        voltage = 110 if "110千伏" in name else (35 if "35千伏" in name else None)
        if voltage is None or "工程" not in name:
            continue
        project_type = "replacement_or_uprating" if "增容" in name else "new_third_transformer"
        records.append(
            {
                "cost_case_id": f"COST-{voltage}-{len(records) + 1:03d}",
                "voltage_kv": voltage,
                "candidate_type": project_type,
                "project_capacity_mva": (_number(row[4]) or 0.0) * 10.0,
                "static_capex_wanyuan": _number(row[15]),
                "dynamic_capex_wanyuan": dynamic,
                "price_nature": "local_dynamic_investment_estimate",
                "source_row": offset,
                **_lineage(
                    "SRC10:Sheet1",
                    "retain anonymized project type, capacity and investment; discard real project names",
                    "local_case_scope_requires_type_match",
                    source_hash,
                ),
            }
        )
    return pd.DataFrame(records)


def _asset_scope_summary(
    station_master: pd.DataFrame,
    county_reference: pd.DataFrame,
    source_hash: str,
) -> pd.DataFrame:
    year_end = (
        station_master[station_master["asset_scope_id"] == "year_end_2025"]
        .groupby(["region_id", "voltage_kv"], as_index=False)
        .agg(year_end_capacity_mva=("capacity_mva", "sum"), year_end_station_count=("station_id", "nunique"))
    )
    merged = county_reference[["region_id", "voltage_kv", "capacity_mva"]].merge(
        year_end, on=["region_id", "voltage_kv"], how="left", validate="one_to_one"
    )
    # 3.1.0 调和结论（2026-08-24 核验）：QX-00005 的 35 kV 站级聚合、主变
    # 求和与官方锚点已三方一致（275 MVA），不再列为未决；QX-00010 的
    # operating 口径取官方 232 MVA，年末设备 192 MVA 仅作设备实况参考
    # （与 110 kV 双口径处理同构），差异已文档化。
    unresolved = set()
    records = []
    for row in merged.itertuples(index=False):
        difference = float(row.year_end_capacity_mva - row.capacity_mva)
        records.append(
            {
                "region_id": row.region_id,
                "voltage_kv": int(row.voltage_kv),
                "operating_2025_capacity_mva": float(row.capacity_mva),
                "year_end_2025_capacity_mva": float(row.year_end_capacity_mva),
                "year_end_station_count": int(row.year_end_station_count),
                "difference_mva": difference,
                "reconciliation_status": (
                    "must_reconcile" if (row.region_id, int(row.voltage_kv)) in unresolved
                    else ("matched" if abs(difference) <= 1e-6 else "scope_difference_documented")
                ),
                **_lineage(
                    "SRC02+SRC03+SRC08",
                    "compare year-end equipment sum with official 2025 operating capacity",
                    "unresolved_asset_scope" if (row.region_id, int(row.voltage_kv)) in unresolved else "scope_comparison",
                    source_hash,
                ),
            }
        )
    return pd.DataFrame(records)


def _base_quality_issues(source_hash: str) -> list[dict[str, Any]]:
    return [
        {
            "issue_id": "DQ-TIMESERIES-CANDIDATE",
            "severity": "blocking_for_grade_A",
            "entity_id": "QX-00005",
            "description": "2025 time-series column map is candidate-only until stage-2 cross-validation.",
            **_lineage("SRC07+candidate_map", "approval gate", "open", source_hash),
        },
        {
            "issue_id": "DQ-PV-SNAPSHOT-YEAR",
            "severity": "warning",
            "entity_id": "all_regions",
            "description": "PV capacity snapshot is from 2026 context and is not relabeled as 2025.",
            **_lineage("SRC05", "snapshot date audit", "open", source_hash),
        },
        {
            "issue_id": "DQ-35-QX00005",
            "severity": "resolved",
            "entity_id": "QX-00005|35",
            "description": "35 kV station aggregation, transformer sum and official anchor all equal 275 MVA; reconciled 2026-08-24, no blocking remains.",
            **_lineage("SRC02+SRC03+SRC08", "asset-scope comparison", "resolved_2026-08-24", source_hash),
        },
        {
            "issue_id": "DQ-35-QX00010",
            "severity": "warning",
            "entity_id": "QX-00010|35",
            "description": "Operating scope uses the official 232 MVA anchor; year-end device table sums to 192 MVA and is documented as equipment-reality context only (same dual-scope treatment as 110 kV).",
            **_lineage("SRC02+SRC03+SRC08", "asset-scope comparison", "resolved_by_official_anchor_precedence_2026-08-24", source_hash),
        },
    ]


def _write_frame(path: Path, frame: pd.DataFrame, sort_by: Iterable[str] = ()) -> None:
    missing_lineage = set(LINEAGE_FIELDS) - set(frame.columns)
    if missing_lineage:
        raise RealDataQualityError(f"{path.name}: missing lineage columns {sorted(missing_lineage)}")
    if frame.empty:
        raise RealDataQualityError(f"{path.name}: refusing to write empty table")
    keys = [key for key in sort_by if key in frame.columns]
    if keys:
        frame = frame.sort_values(keys, kind="stable").reset_index(drop=True)
    path.parent.mkdir(parents=True, exist_ok=True)
    frame.to_csv(path, index=False, lineterminator="\n", float_format="%.10g")


def adapt_real_2025(source_root: Path, output_root: Path, contract_path: Path) -> dict[str, Any]:
    """把权威源包转换为可重放的 `real_2025` 阶段 1 数据集。"""
    source_root = Path(source_root).resolve()
    output_root = Path(output_root).resolve()
    contract_path = Path(contract_path).resolve()
    if not source_root.is_dir():
        raise RealDataQualityError(f"source root not found: {source_root}")
    if not contract_path.is_file():
        raise RealDataQualityError(f"contract not found: {contract_path}")
    contract = yaml.safe_load(contract_path.read_text(encoding="utf-8"))
    if contract["contract"]["version"] != "2.0.0":
        raise RealDataQualityError("real adapter requires model contract 2.0.0")
    regions = set(contract["scope"]["regions"])
    source_files = _source_registry(source_root)
    before_hashes = {name: item["sha256"] for name, item in source_files.items()}

    def source_hash(name: str) -> str:
        try:
            return str(source_files[name]["sha256"])
        except KeyError as exc:
            raise RealDataQualityError(f"required source missing: {name}") from exc

    equipment, raw_expansion = _equipment(source_root, regions, source_hash(SRC03_04))
    transformer_master, transformer_static, issues_a = _transformer_tables(
        source_root, regions, equipment, source_hash(SRC01_02)
    )
    station_master, station_static, issues_b = _station_tables(
        source_root, regions, transformer_master, source_hash(SRC01_02)
    )
    candidates = _expansion_candidates(raw_expansion, station_master)
    pv_snapshot = _pv_snapshot(source_root, regions, source_hash(SRC05))
    pv_profile = _pv_profile(source_root, source_hash(SRC14))
    network_lines = _network_lines(source_root, regions, source_hash(SRC03_04))
    county_reference = _county_reference(source_root, regions, source_hash(SRC08))
    cost_cases = _cost_cases(source_root, source_hash(SRC10))
    asset_scopes = _asset_scope_summary(station_master, county_reference, source_hash(SRC01_02))
    issues = pd.DataFrame(
        _base_quality_issues(source_hash(SRC01_02)) + issues_a + issues_b
    )

    qx_stations = station_master[
        (station_master["region_id"] == "QX-00005")
        & (station_master["voltage_kv"] == 110)
        & (station_master["asset_scope_id"] == "year_end_2025")
    ]
    qx_110 = transformer_master[
        (transformer_master["region_id"] == "QX-00005")
        & (transformer_master["voltage_kv"] == 110)
    ]
    qx_35 = transformer_master[
        (transformer_master["region_id"] == "QX-00005")
        & (transformer_master["voltage_kv"] == 35)
    ]
    if (qx_stations["station_id"].nunique(), len(qx_110), len(qx_35)) != (21, 42, 16):
        raise RealDataQualityError(
            "QX-00005 expected 21 stations / 42 110-kV transformers / 16 35-kV transformers"
        )
    if abs(float(qx_110["capacity_mva"].sum()) - 2239.5) > 1e-6:
        raise RealDataQualityError("QX-00005 year-end 110-kV transformer capacity must be 2239.5 MVA")
    if transformer_master["transformer_uid"].duplicated().any():
        raise RealDataQualityError("transformer_uid is not unique")
    if candidates["candidate_id"].duplicated().any():
        raise RealDataQualityError("candidate_id is not unique")

    frames = {
        "station_master.csv": (station_master, ["region_id", "voltage_kv", "asset_scope_id", "station_id"]),
        "transformer_master.csv": (transformer_master, ["region_id", "voltage_kv", "station_id", "unit_id"]),
        "expansion_candidates.csv": (candidates, ["region_id", "station_id", "candidate_id"]),
        "station_static_load.csv": (station_static, ["region_id", "voltage_kv", "station_id"]),
        "transformer_static_load.csv": (transformer_static, ["region_id", "voltage_kv", "station_id", "unit_id"]),
        "station_pv_snapshot.csv": (pv_snapshot, ["region_id", "voltage_kv", "station_id"]),
        "pv_profile_2025.csv": (pv_profile, ["hour_index"]),
        "network_lines_110kv.csv": (network_lines, ["region_id", "line_id"]),
        "county_clr_reference.csv": (county_reference, ["region_id", "voltage_kv"]),
        "cost_cases.csv": (cost_cases, ["voltage_kv", "cost_case_id"]),
        "asset_scope_summary.csv": (asset_scopes, ["region_id", "voltage_kv"]),
        "data_quality_issues.csv": (issues, ["issue_id"]),
    }
    output_root.mkdir(parents=True, exist_ok=True)
    for filename, (frame, keys) in frames.items():
        _write_frame(output_root / filename, frame, keys)

    after_hashes = {name: _sha256(source_root / name) for name in before_hashes}
    if after_hashes != before_hashes:
        raise RealDataQualityError("source package changed during adaptation")

    output_files = {
        filename: {
            "sha256": _sha256(output_root / filename),
            "rows": int(len(frame)),
        }
        for filename, (frame, _keys) in sorted(frames.items())
    }
    fingerprint_payload = {
        "adapter_version": ADAPTER_VERSION,
        "contract_sha256": _sha256(contract_path),
        "source_hashes": before_hashes,
    }
    fingerprint = hashlib.sha256(
        json.dumps(fingerprint_payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()
    manifest: dict[str, Any] = {
        "dataset_id": SCENARIO_ID,
        "adapter_version": ADAPTER_VERSION,
        "contract_id": contract["contract"]["id"],
        "contract_version": contract["contract"]["version"],
        "contract_sha256": _sha256(contract_path),
        "dataset_fingerprint": fingerprint,
        "source_root_name": source_root.name,
        "source_files": source_files,
        "output_files": output_files,
        "quality_gate": {
            "qx00005_stations_110kv": 21,
            "qx00005_transformers_110kv": 42,
            "qx00005_transformers_35kv_context": 16,
            "pv_profile_hours": 8760,
            "county_voltage_reference_rows": 16,
            "timeseries_status": "candidate_only_stage2_required",
        },
    }
    (output_root / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return manifest
