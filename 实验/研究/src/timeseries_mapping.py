"""SRC07 2025 年 58 列主变时序映射交叉验证与质量门禁。"""
from __future__ import annotations

import hashlib
import json
import math
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import yaml
from openpyxl import load_workbook


SOURCE_VERSION = "电网建模数据_Agent整合版_V1.2"
HOURLY_FILE = "邳州主变负载率.xlsx"
PV_FILE = "光伏装机.xlsx"


class TimeseriesMappingError(ValueError):
    """时序映射结构或质量门禁失败。"""


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _number(value: Any) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def _voltage(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value).lower().replace("kv", "").replace("千伏", "").strip()
    try:
        return int(float(text))
    except ValueError:
        return None


def _unit(value: Any) -> str | None:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip().replace("＃", "#")
    return text if text.startswith("#") else f"#{text}"


def _parse_candidate(label: str) -> tuple[str, str, int]:
    parts = str(label).split("/")
    if len(parts) != 3:
        raise TimeseriesMappingError(f"invalid candidate label: {label!r}")
    station_id, unit_id, voltage_label = parts
    voltage = _voltage(voltage_label)
    if voltage not in {35, 110}:
        raise TimeseriesMappingError(f"invalid candidate voltage: {label!r}")
    normalized_unit = _unit(unit_id)
    if normalized_unit is None:
        raise TimeseriesMappingError(f"invalid candidate unit: {label!r}")
    return station_id, normalized_unit, voltage


def _read_hourly_sheet(path: Path, sheet_name: str) -> tuple[pd.DataFrame, pd.Series]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        timestamps: list[pd.Timestamp] = []
        values: list[list[float]] = []
        source_rows: list[int] = []
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=3, max_row=sheet.max_row, max_col=59, values_only=True),
            start=3,
        ):
            timestamp = pd.to_datetime(row[0], errors="coerce")
            if pd.isna(timestamp):
                continue
            timestamps.append(timestamp)
            values.append([
                float(value) if _number(value) is not None else float("nan")
                for value in row[1:59]
            ])
            source_rows.append(row_number)
    finally:
        workbook.close()
    index = pd.DatetimeIndex(timestamps, name="timestamp")
    frame = pd.DataFrame(values, index=index, columns=range(1, 59), dtype=float)
    return frame, pd.Series(source_rows, index=index, name="source_row")


def _read_snapshot_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[dict[str, Any]] = []
    try:
        sheet = workbook["主变（跨区合并）1"]
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, max_row=521, max_col=10, values_only=True),
            start=2,
        ):
            if row[1] != "QX-00005" or _voltage(row[3]) not in {35, 110}:
                continue
            timestamp = pd.to_datetime(row[4], errors="coerce")
            net_load = _number(row[5])
            capacity = _number(row[6])
            if pd.isna(timestamp) or net_load is None or capacity is None:
                continue
            records.append(
                {
                    "station_id": str(row[2]),
                    "voltage_kv": int(_voltage(row[3]) or 0),
                    "timestamp": timestamp,
                    "net_load_mw": net_load,
                    "capacity_mva": capacity,
                    "source_row": row_number,
                }
            )
    finally:
        workbook.close()
    return records


def _snapshot_error(
    series_2026: pd.Series,
    snapshots: list[dict[str, Any]],
    station_id: str,
    voltage_kv: int,
    capacity_mva: float,
) -> tuple[float | None, int]:
    matches = [
        row
        for row in snapshots
        if row["station_id"] == station_id
        and row["voltage_kv"] == voltage_kv
        and abs(row["capacity_mva"] - capacity_mva) <= 1e-6
    ]
    errors = []
    valid = series_2026.dropna()
    for row in matches:
        location = valid.index.get_indexer(
            [row["timestamp"]], method="nearest", tolerance=pd.Timedelta(hours=1)
        )[0]
        if location < 0:
            continue
        observed = float(valid.iloc[location])
        scale = max(capacity_mva, 10.0)
        errors.append(abs(observed - row["net_load_mw"]) / scale)
    return (min(errors) if errors else None), len(errors)


def _isolated_imputation(raw: pd.Series) -> tuple[pd.Series, np.ndarray]:
    corrected = raw.copy()
    imputed = np.zeros(len(raw), dtype=bool)
    missing = raw.isna().to_numpy()
    for index in np.flatnonzero(missing):
        if (
            0 < index < len(raw) - 1
            and not missing[index - 1]
            and not missing[index + 1]
        ):
            corrected.iloc[index] = (float(raw.iloc[index - 1]) + float(raw.iloc[index + 1])) / 2.0
            imputed[index] = True
    return corrected, imputed


def _mapping_status(
    score: float,
    max_error: float | None,
    min_error: float | None,
    snapshot_error: float | None,
    nonmissing_hours: int,
    capacity_match: bool,
) -> tuple[str, str, str]:
    if not capacity_match:
        return "rejected", "low", "candidate transformer capacity/key does not match authoritative master"
    if score >= 999 or nonmissing_hours < 8000:
        return "rejected", "low", "insufficient 2025 coverage or no finite extrema match"
    errors = [value for value in (max_error, min_error) if value is not None]
    extrema_error = max(errors) if errors else float("inf")
    if score <= 0.10 and extrema_error <= 0.15 and snapshot_error is not None and snapshot_error <= 0.08:
        confidence = "high"
    elif score <= 0.25 and extrema_error <= 0.30 and snapshot_error is not None and snapshot_error <= 0.15:
        confidence = "medium"
    else:
        confidence = "low"
    return (
        "conditional",
        confidence,
        "algorithmic cross-check only; source-owner approval is still required for formal 8760 use",
    )


def _mapping_lineage(hourly_hash: str, quality_flag: str) -> dict[str, str]:
    return {
        "source_ref": "SRC07:Sheet3+SRC02+SRC05+candidate_map",
        "source_version": SOURCE_VERSION,
        "transformation": "cross-check candidate against 2025 extrema, rated capacity and 2026 snapshot",
        "scenario_id": "real_2025_timeseries_review",
        "quality_flag": quality_flag,
        "source_sha256": hourly_hash,
    }


def review_real_timeseries_mapping(
    source_root: Path,
    candidate_map_path: Path,
    processed_root: Path,
    contract_path: Path,
) -> dict[str, Any]:
    """交叉验证候选映射；未获外部审批时绝不自动输出 A 级许可。"""
    source_root = Path(source_root).resolve()
    candidate_map_path = Path(candidate_map_path).resolve()
    processed_root = Path(processed_root).resolve()
    contract_path = Path(contract_path).resolve()
    contract = yaml.safe_load(contract_path.read_text(encoding="utf-8"))
    current_status = contract["data"]["timeseries"]["current_status"]
    if current_status not in {
        "candidate_only",
        "approved_for_2025_qx00005_110kv_operating_scope",
    }:
        raise TimeseriesMappingError("unexpected contract time-series status")
    required_processed = ["transformer_master.csv", "manifest.json"]
    for filename in required_processed:
        if not (processed_root / filename).is_file():
            raise TimeseriesMappingError(f"run real-data adapter first: missing {filename}")

    candidate = pd.read_csv(candidate_map_path)
    expected_candidate_cols = {
        "series_column_1_based",
        "candidate",
        "score",
        "max_error_fraction",
        "min_error_fraction",
    }
    if len(candidate) != 58 or not expected_candidate_cols <= set(candidate.columns):
        raise TimeseriesMappingError("candidate map must contain 58 rows and diagnostic fields")
    if candidate["series_column_1_based"].duplicated().any() or candidate["candidate"].duplicated().any():
        raise TimeseriesMappingError("candidate map columns and transformer labels must be unique")

    transformer_master = pd.read_csv(processed_root / "transformer_master.csv")
    qx_master = transformer_master[transformer_master["region_id"] == "QX-00005"].copy()
    master_by_key = {
        (row.station_id, row.unit_id, int(row.voltage_kv)): row
        for row in qx_master.itertuples(index=False)
    }

    hourly_path = source_root / HOURLY_FILE
    hourly_hash = _sha256(hourly_path)
    pv_hash = _sha256(source_root / PV_FILE)
    candidate_hash = _sha256(candidate_map_path)
    sheet3, source_rows = _read_hourly_sheet(hourly_path, "Sheet3")
    year2025 = sheet3[sheet3.index.year == 2025].copy()
    rows2025 = source_rows[source_rows.index.year == 2025].copy()
    if len(year2025) != 8760 or year2025.index.nunique() != 8760:
        raise TimeseriesMappingError("Sheet3 must contain exactly 8760 unique 2025 timestamps")
    expected_axis = pd.date_range("2025-01-01", periods=8760, freq="h")
    if not year2025.index.equals(expected_axis):
        raise TimeseriesMappingError("Sheet3 2025 timestamps are not continuous hourly values")

    sheet4, _ = _read_hourly_sheet(hourly_path, "Sheet4")
    snapshots = _read_snapshot_rows(source_root / PV_FILE)
    map_records: list[dict[str, Any]] = []
    for row in candidate.sort_values("series_column_1_based").itertuples(index=False):
        column = int(row.series_column_1_based)
        station_id, unit_id, voltage_kv = _parse_candidate(row.candidate)
        master = master_by_key.get((station_id, unit_id, voltage_kv))
        capacity = float(master.capacity_mva) if master is not None else float("nan")
        series = year2025[column]
        nonmissing = int(series.notna().sum())
        snapshot_error, snapshot_count = _snapshot_error(
            sheet4[column], snapshots, station_id, voltage_kv, capacity
        ) if master is not None else (None, 0)
        max_error = _number(row.max_error_fraction)
        min_error = _number(row.min_error_fraction)
        score = float(row.score)
        capacity_match = master is not None and math.isfinite(capacity) and capacity > 0
        status, confidence, reason = _mapping_status(
            score, max_error, min_error, snapshot_error, nonmissing, capacity_match
        )
        map_records.append(
            {
                "series_column_1_based": column,
                "transformer_uid": (
                    master.transformer_uid if master is not None
                    else f"QX-00005|{voltage_kv}|{station_id}|{unit_id}"
                ),
                "region_id": "QX-00005",
                "voltage_kv": voltage_kv,
                "station_id": station_id,
                "unit_id": unit_id,
                "capacity_mva": capacity,
                "capacity_match": bool(capacity_match),
                "candidate_score": score,
                "candidate_best_margin": _number(getattr(row, "independent_best_margin", None)),
                "max_error_fraction": max_error,
                "min_error_fraction": min_error,
                "snapshot_2026_error_fraction": snapshot_error,
                "snapshot_2026_match_count": snapshot_count,
                "nonmissing_hours_2025": nonmissing,
                "missing_hours_2025": int(8760 - nonmissing),
                "observed_max_2025_mw": float(series.max()) if nonmissing else None,
                "observed_min_2025_mw": float(series.min()) if nonmissing else None,
                "approval_status": status,
                "confidence": confidence,
                "approval_reason": reason,
                "formal_use_allowed": False,
                "approval_authority": "pending_source_owner",
                "candidate_map_sha256": candidate_hash,
                "pv_snapshot_sha256": pv_hash,
                **_mapping_lineage(hourly_hash, f"mapping_{status}_{confidence}"),
            }
        )
    mapping = pd.DataFrame(map_records).sort_values("series_column_1_based").reset_index(drop=True)
    if mapping["transformer_uid"].duplicated().any():
        raise TimeseriesMappingError("reviewed mapping is not one-to-one")
    split = mapping.groupby("voltage_kv").size().to_dict()
    if split != {35: 16, 110: 42}:
        raise TimeseriesMappingError(f"mapping split must be 42/16, got {split}")

    mapping_path = processed_root / "timeseries_column_map_2025.csv"
    mapping.to_csv(mapping_path, index=False, lineterminator="\n", float_format="%.10g")

    hourly_frames: list[pd.DataFrame] = []
    quality_issues: list[dict[str, Any]] = []
    for map_row in mapping.itertuples(index=False):
        column = int(map_row.series_column_1_based)
        raw = year2025[column].copy()
        corrected, imputed = _isolated_imputation(raw)
        missing_unfilled = corrected.isna().to_numpy()
        outlier = corrected.abs().gt(2.0 * float(map_row.capacity_mva)).fillna(False).to_numpy()
        corrected.iloc[np.flatnonzero(outlier)] = np.nan
        flags = np.full(8760, "source_value", dtype=object)
        flags[imputed] = "imputed_linear_isolated"
        flags[missing_unfilled] = "missing_not_imputed"
        flags[outlier] = "outlier_removed_not_imputed"
        hourly_frames.append(
            pd.DataFrame(
                {
                    "timestamp": year2025.index.strftime("%Y-%m-%d %H:%M:%S"),
                    "source_row": rows2025.to_numpy(),
                    "series_column_1_based": column,
                    "transformer_uid": map_row.transformer_uid,
                    "region_id": "QX-00005",
                    "voltage_kv": int(map_row.voltage_kv),
                    "station_id": map_row.station_id,
                    "unit_id": map_row.unit_id,
                    "net_load_mw_raw": raw.to_numpy(),
                    "net_load_mw": corrected.to_numpy(),
                    "point_quality_flag": flags,
                    "mapping_approval_status": map_row.approval_status,
                    "formal_use_allowed": False,
                    "source_ref": "SRC07:Sheet3",
                    "source_version": SOURCE_VERSION,
                    "transformation": "retain raw value; only interpolate isolated internal one-hour gaps; remove rating outliers",
                    "scenario_id": "real_2025_timeseries_review",
                    "quality_flag": flags,
                    "source_sha256": hourly_hash,
                }
            )
        )
        if map_row.approval_status == "rejected" or int(map_row.missing_hours_2025) > 0:
            quality_issues.append(
                {
                    "issue_id": f"TS-{column:02d}",
                    "series_column_1_based": column,
                    "transformer_uid": map_row.transformer_uid,
                    "approval_status": map_row.approval_status,
                    "missing_hours_2025": int(map_row.missing_hours_2025),
                    "description": map_row.approval_reason,
                    **_mapping_lineage(hourly_hash, f"mapping_{map_row.approval_status}"),
                }
            )
    hourly = pd.concat(hourly_frames, ignore_index=True)
    hourly_path_out = processed_root / "transformer_hourly_2025.csv.gz"
    hourly.to_csv(
        hourly_path_out,
        index=False,
        float_format="%.10g",
        compression={"method": "gzip", "compresslevel": 6, "mtime": 0},
    )
    issue_frame = pd.DataFrame(quality_issues)
    issue_path = processed_root / "timeseries_quality_issues.csv"
    issue_frame.sort_values("series_column_1_based").to_csv(
        issue_path, index=False, lineterminator="\n", float_format="%.10g"
    )

    mapping_fingerprint = _sha256(mapping_path)
    approved_110 = int(
        ((mapping["voltage_kv"] == 110) & (mapping["approval_status"] == "approved")).sum()
    )
    result = {
        "mapping_fingerprint": mapping_fingerprint,
        "total_columns": 58,
        "conditional_columns": int((mapping["approval_status"] == "conditional").sum()),
        "rejected_columns": int((mapping["approval_status"] == "rejected").sum()),
        "approved_110kv_columns": approved_110,
        "grade_a_ready": approved_110 == 42,
        "formal_hourly_use_allowed": False,
    }
    manifest_path = processed_root / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["timeseries_review"] = result
    manifest["output_files"][mapping_path.name] = {
        "sha256": mapping_fingerprint,
        "rows": 58,
    }
    manifest["output_files"][hourly_path_out.name] = {
        "sha256": _sha256(hourly_path_out),
        "rows": int(len(hourly)),
    }
    manifest["output_files"][issue_path.name] = {
        "sha256": _sha256(issue_path),
        "rows": int(len(issue_frame)),
    }
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return result
