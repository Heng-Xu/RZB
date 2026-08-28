"""v3 年度资产白名单、跨年时序质量和项目负责人审批接口。"""
from __future__ import annotations

import hashlib
import json
import math
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
import yaml
from openpyxl import load_workbook

from src.real_data_adapter import (
    SRC05,
    SRC08,
    SRC14,
    SRC01_02,
    SRC03_04,
    SOURCE_VERSION,
    _asset_scope_summary,
    _cost_cases,
    _equipment,
    _expansion_candidates,
    _network_lines,
    _pv_profile,
    _pv_snapshot,
    _rows,
    _sha256 as _adapter_sha256,
    _station_tables,
    _transformer_tables,
)
from src.timeseries_mapping import (
    _isolated_imputation,
    _mapping_status,
    _parse_candidate,
    _read_hourly_sheet,
    _read_snapshot_rows,
    _snapshot_error,
)


DATASET_ID = "real_2021_2025"
ADAPTER_VERSION = "3.0.0"
LINEAGE_FIELDS = (
    "source_ref",
    "source_version",
    "transformation",
    "scenario_id",
    "quality_flag",
    "source_sha256",
)
HOURLY_FILE = "邳州主变负载率.xlsx"
CANDIDATE_FILE = "results/real_data_audit/timeseries_column_map_2025_candidate.csv"
YEAR_SHEETS = {2022: "Sheet1", 2023: "Sheet2", 2024: "Sheet3", 2025: "Sheet3", 2026: "Sheet4"}
YEAR_HOURS = {2022: 8760, 2023: 8760, 2024: 8784, 2025: 8760, 2026: 4992}


class AnnualAssetScopeError(ValueError):
    """年度资产范围、时序门禁或来源血缘不满足 v3 约束。"""


def _sha256(path: Path) -> str:
    return _adapter_sha256(path)


def _lineage(
    source_ref: str,
    transformation: str,
    quality_flag: str,
    source_sha256: str,
    scenario_id: str = DATASET_ID,
) -> dict[str, str]:
    return {
        "source_ref": source_ref,
        "source_version": SOURCE_VERSION,
        "transformation": transformation,
        "scenario_id": scenario_id,
        "quality_flag": quality_flag,
        "source_sha256": source_sha256,
    }


def _write_frame(path: Path, frame: pd.DataFrame, sort_by: Iterable[str] = ()) -> None:
    if frame.empty:
        raise AnnualAssetScopeError(f"{path.name}: refusing to write empty table")
    missing = set(LINEAGE_FIELDS) - set(frame.columns)
    if missing:
        raise AnnualAssetScopeError(f"{path.name}: missing lineage columns {sorted(missing)}")
    keys = [key for key in sort_by if key in frame.columns]
    if keys:
        frame = frame.sort_values(keys, kind="stable").reset_index(drop=True)
    path.parent.mkdir(parents=True, exist_ok=True)
    compression = "gzip" if path.suffix == ".gz" else None
    frame.to_csv(
        path,
        index=False,
        lineterminator="\n",
        float_format="%.10g",
        compression=compression,
    )


def _source_headers(path: Path, sheet_name: str) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        row = next(workbook[sheet_name].iter_rows(min_row=1, max_row=1, max_col=59, values_only=True))
    finally:
        workbook.close()
    headers: list[str] = []
    for value in row[1:59]:
        text = "" if value is None else str(value).strip()
        headers.append(text)
    if len(headers) != 58:
        raise AnnualAssetScopeError(f"{sheet_name}: expected 58 source headers")
    return headers


def _annual_reference(source_root: Path, regions: set[str], source_hash: str) -> pd.DataFrame:
    """读取 SRC08 五年官方容量、峰值和容载比锚点，不反推设备历史。"""
    rows = _rows(source_root / SRC08, "Sheet1", 4, 80, 17)
    # SRC08 的五年块从第 C、F、I、L、O 列开始，单位为万 kVA/万 kW。
    year_columns = {2021: 2, 2022: 5, 2023: 8, 2024: 11, 2025: 14}
    records: list[dict[str, Any]] = []
    current_region: str | None = None
    for offset, row in enumerate(rows, start=4):
        if row[0] is not None:
            current_region = str(row[0]).strip()
        voltage_text = row[1]
        if current_region not in regions or voltage_text is None:
            continue
        voltage = 110 if "110" in str(voltage_text) else (35 if "35" in str(voltage_text) else None)
        if voltage is None:
            continue
        for year, capacity_col in year_columns.items():
            capacity = row[capacity_col]
            peak = row[capacity_col + 1]
            clr = row[capacity_col + 2]
            try:
                values = [float(capacity) * 10.0, float(peak) * 10.0, float(clr)]
            except (TypeError, ValueError):
                continue
            if not all(math.isfinite(value) for value in values):
                continue
            records.append(
                {
                    "year": year,
                    "region_id": current_region,
                    "voltage_kv": voltage,
                    "official_capacity_mva": values[0],
                    "official_positive_peak_mw": values[1],
                    "official_clr": values[2],
                    "source_row": offset,
                    **_lineage(
                        "SRC08:Sheet1",
                        "convert annual official 10000 kVA/kW anchors to MVA/MW; retain as total-volume anchor",
                        "official_annual_anchor",
                        source_hash,
                    ),
                }
            )
    frame = pd.DataFrame(records)
    expected = len(regions) * 2 * 5
    if len(frame) != expected:
        raise AnnualAssetScopeError(f"SRC08 expected {expected} annual rows, got {len(frame)}")
    return frame


def build_annual_asset_whitelist(
    transformer_master: pd.DataFrame,
    annual_reference: pd.DataFrame,
    source_hash: str,
) -> pd.DataFrame:
    """生成年度白名单；历史设备范围未知时显式降级，不按投运年回溯。"""
    records: list[dict[str, Any]] = []
    for row in transformer_master.itertuples(index=False):
        uid = str(row.transformer_uid)
        is_qx_late = (
            row.region_id == "QX-00005"
            and int(row.voltage_kv) == 110
            and str(row.station_id) == "BDZ-00056"
        )
        for year in range(2021, 2026):
            if year == 2025 and is_qx_late:
                scopes = [("year_end_2025", True), ("operating_2025", False)]
            elif year == 2025:
                scopes = [("year_end_2025", True), ("operating_2025", True)]
            else:
                scopes = [("historical_reference_only", False)]
            for scope_id, in_scope in scopes:
                quality = (
                    "year_end_only_not_in_operating_scope"
                    if scope_id == "year_end_2025" and is_qx_late
                    else (
                        "operating_whitelist_2025"
                        if scope_id == "operating_2025"
                        else "historical_device_scope_not_closed"
                    )
                )
                records.append(
                    {
                        "year": year,
                        "region_id": row.region_id,
                        "voltage_kv": int(row.voltage_kv),
                        "station_id": row.station_id,
                        "transformer_uid": uid,
                        "capacity_mva": float(row.capacity_mva),
                        "asset_scope_id": scope_id,
                        "in_annual_operating_whitelist": bool(in_scope),
                        "asset_presence_status": "known_current_equipment" if year == 2025 else "not_closed",
                        "scope_basis": "SRC02+SRC03+2025_operating_rule" if year == 2025 else "official_annual_anchor_without_device_backfill",
                        **_lineage(
                            "SRC02:主变1+SRC03:设备明细+SRC08",
                            "construct annual whitelist without backfilling historical device presence from current commissioning years",
                            quality,
                            source_hash,
                        ),
                    }
                )
    return pd.DataFrame(records)


def build_annual_reconciliation(
    whitelist: pd.DataFrame,
    annual_reference: pd.DataFrame,
    source_hash: str,
) -> pd.DataFrame:
    current = (
        whitelist[whitelist["asset_scope_id"] == "year_end_2025"]
        .groupby(["region_id", "voltage_kv"], as_index=False)
        .agg(device_table_capacity_mva=("capacity_mva", "sum"), device_table_transformer_count=("transformer_uid", "nunique"))
    )
    operating = (
        whitelist[whitelist["asset_scope_id"] == "operating_2025"]
        .groupby(["region_id", "voltage_kv"], as_index=False)
        .agg(operating_2025_capacity_mva=("capacity_mva", "sum"), operating_2025_transformer_count=("transformer_uid", "nunique"))
    )
    frame = annual_reference.merge(current, on=["region_id", "voltage_kv"], how="left", validate="many_to_one")
    frame = frame.merge(operating, on=["region_id", "voltage_kv"], how="left", validate="many_to_one")
    frame["capacity_difference_mva"] = frame["device_table_capacity_mva"] - frame["official_capacity_mva"]
    frame["reconciliation_status"] = np.select(
        [
            frame["year"].lt(2025),
            frame["year"].eq(2025) & frame["region_id"].eq("QX-00005") & frame["voltage_kv"].eq(110) & frame["capacity_difference_mva"].abs().lt(1e-6),
            frame["capacity_difference_mva"].abs().lt(1e-6),
        ],
        ["historical_device_scope_not_closed", "operating_anchor_reconciled", "scope_difference_documented"],
        default="must_reconcile",
    )
    frame["quality_flag"] = frame["reconciliation_status"].map(
        lambda value: "official_anchor_and_device_scope_reconciled" if value == "operating_anchor_reconciled" else value
    )
    frame["source_ref"] = "SRC08:Sheet1+SRC02+SRC03"
    frame["source_version"] = SOURCE_VERSION
    frame["transformation"] = "compare annual official total anchor with year-end device table; do not infer historical devices"
    frame["scenario_id"] = DATASET_ID
    frame["source_sha256"] = source_hash
    return frame


def build_actual_asset_actions(
    reconciliation: pd.DataFrame,
    transformer_master: pd.DataFrame,
    source_hash: str,
) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    # 2025 年末新增设备只记录事实范围，成本没有设备级闭合证据时保持“未识别”。
    late = transformer_master[
        (transformer_master["region_id"] == "QX-00005")
        & (transformer_master["voltage_kv"] == 110)
        & (transformer_master["station_id"] == "BDZ-00056")
    ]
    for row in late.itertuples(index=False):
        records.append(
            {
                "action_id": f"ACT-2025-{row.transformer_uid}",
                "year": 2025,
                "region_id": row.region_id,
                "voltage_kv": int(row.voltage_kv),
                "station_id": row.station_id,
                "transformer_uids": row.transformer_uid,
                "action_type": "new_in_service_scope",
                "capacity_delta_mva": float(row.capacity_mva),
                "action_status": "fact_scope_only",
                "capex_wanyuan_2025": "未识别",
                "eac_wanyuan_per_year": "未识别",
                "cost_closure_status": "device_cost_not_closed",
                "action_basis": "year_end_2025_only_equipment_scope",
                **_lineage(
                    "SRC02:主变1+SRC03:设备明细",
                    "record year-end new equipment without inventing historical action cost",
                    "actual_action_cost_unidentified",
                    source_hash,
                ),
            }
        )
    for row in reconciliation.itertuples(index=False):
        if int(row.year) < 2022 or row.reconciliation_status == "operating_anchor_reconciled":
            continue
        records.append(
            {
                "action_id": f"ACT-UNRESOLVED-{row.year}-{row.region_id}-{int(row.voltage_kv)}",
                "year": int(row.year),
                "region_id": row.region_id,
                "voltage_kv": int(row.voltage_kv),
                "station_id": "",
                "transformer_uids": "",
                "action_type": "actual_change_not_closed",
                "capacity_delta_mva": "未识别",
                "action_status": "not_closed",
                "capex_wanyuan_2025": "未识别",
                "eac_wanyuan_per_year": "未识别",
                "cost_closure_status": "device_action_not_closed",
                "action_basis": row.reconciliation_status,
                **_lineage(
                    "SRC08:Sheet1+SRC02+SRC03",
                    "record official/device scope discrepancy without converting it to a device action",
                    "actual_action_not_identified",
                    source_hash,
                ),
            }
        )
    return pd.DataFrame(records)


def _candidate_records(
    source_root: Path,
    processed_root: Path,
    candidate_path: Path,
    source_hash: str,
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    candidate = pd.read_csv(candidate_path)
    required = {"series_column_1_based", "candidate", "score", "max_error_fraction", "min_error_fraction"}
    if len(candidate) != 58 or not required <= set(candidate.columns):
        raise AnnualAssetScopeError("candidate map must contain 58 rows and diagnostic fields")
    if candidate["series_column_1_based"].duplicated().any() or candidate["candidate"].duplicated().any():
        raise AnnualAssetScopeError("candidate map columns and transformer labels must be unique")
    master = pd.read_csv(processed_root / "transformer_master.csv")
    qx_master = master[master["region_id"] == "QX-00005"]
    master_by_key = {(r.station_id, r.unit_id, int(r.voltage_kv)): r for r in qx_master.itertuples(index=False)}
    hourly_path = source_root / HOURLY_FILE
    sheet3, _rows_2025 = _read_hourly_sheet(hourly_path, "Sheet3")
    year2025 = sheet3[sheet3.index.year == 2025]
    sheet4, _ = _read_hourly_sheet(hourly_path, "Sheet4")
    snapshots = _read_snapshot_rows(source_root / SRC05)
    headers = _source_headers(hourly_path, "Sheet3")
    candidate_hash = _sha256(candidate_path)
    records: list[dict[str, Any]] = []
    for item in candidate.sort_values("series_column_1_based").itertuples(index=False):
        column = int(item.series_column_1_based)
        station_id, unit_id, voltage = _parse_candidate(item.candidate)
        master_row = master_by_key.get((station_id, unit_id, voltage))
        capacity = float(master_row.capacity_mva) if master_row is not None else float("nan")
        series = year2025[column]
        snapshot_error, snapshot_count = _snapshot_error(sheet4[column], snapshots, station_id, voltage, capacity) if master_row is not None else (None, 0)
        status, confidence, reason = _mapping_status(
            float(item.score),
            float(item.max_error_fraction) if pd.notna(item.max_error_fraction) else None,
            float(item.min_error_fraction) if pd.notna(item.min_error_fraction) else None,
            snapshot_error,
            int(series.notna().sum()),
            master_row is not None and math.isfinite(capacity) and capacity > 0,
        )
        target_uid = master_row.transformer_uid if master_row is not None else f"QX-00005|{voltage}|{station_id}|{unit_id}"
        records.append(
            {
                "series_column_1_based": column,
                "source_header_station_id": headers[column - 1],
                "transformer_uid": target_uid,
                "region_id": "QX-00005",
                "voltage_kv": voltage,
                "station_id": station_id,
                "unit_id": unit_id,
                "capacity_mva": capacity,
                "candidate_score": float(item.score),
                "max_error_fraction": float(item.max_error_fraction) if pd.notna(item.max_error_fraction) else np.nan,
                "min_error_fraction": float(item.min_error_fraction) if pd.notna(item.min_error_fraction) else np.nan,
                "snapshot_2026_error_fraction": snapshot_error,
                "snapshot_2026_match_count": snapshot_count,
                "nonmissing_hours_2025": int(series.notna().sum()),
                "approval_status": status,
                "confidence": confidence,
                "approval_reason": reason,
                "formal_use_allowed": False,
                "candidate_map_sha256": candidate_hash,
                **_lineage(
                    "SRC07:Sheet3+SRC02+SRC05+candidate_map",
                    "cross-check candidate against 2025 extrema, capacity and 2026 snapshot; retain source header",
                    f"mapping_{status}_{confidence}",
                    source_hash,
                    scenario_id=DATASET_ID,
                ),
            }
        )
    mapping = pd.DataFrame(records)
    if mapping["transformer_uid"].duplicated().any():
        raise AnnualAssetScopeError("v3 mapping is not one-to-one")
    split = mapping.groupby("voltage_kv").size().to_dict()
    if split != {35: 16, 110: 42}:
        raise AnnualAssetScopeError(f"v3 mapping split must be 42/16, got {split}")
    approval = mapping[[
        "series_column_1_based", "source_header_station_id", "transformer_uid", "approval_status",
        "candidate_map_sha256",
    ]].copy()
    approval["approval_authority"] = "project_owner"
    approval["approval_date"] = np.nan
    approval["approval_basis"] = "candidate_cross_check_pending_project_owner_approval"
    approval["source_sha256"] = source_hash
    approval["source_ref"] = "SRC07:Sheet3+candidate_map"
    approval["source_version"] = SOURCE_VERSION
    approval["transformation"] = "preserve source header and candidate target; no self-approval"
    approval["scenario_id"] = DATASET_ID
    approval["quality_flag"] = "pending_project_owner_approval"
    return mapping, approval, {
        "candidate_map_sha256": candidate_hash,
        "source_sha256": source_hash,
        "total_columns": 58,
        "conditional_columns": int(mapping["approval_status"].eq("conditional").sum()),
        "rejected_columns": int(mapping["approval_status"].eq("rejected").sum()),
    }


def _build_hourly_year(
    source_root: Path,
    mapping: pd.DataFrame,
    year: int,
    source_hash: str,
) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    sheet = YEAR_SHEETS[year]
    frame, source_rows = _read_hourly_sheet(source_root / HOURLY_FILE, sheet)
    frame = frame[frame.index.year == year].copy()
    source_rows = source_rows[source_rows.index.year == year].copy()
    expected = YEAR_HOURS[year]
    if len(frame) != expected or frame.index.nunique() != expected:
        raise AnnualAssetScopeError(f"{sheet} {year}: expected {expected} hourly rows")
    records: list[pd.DataFrame] = []
    issues: list[dict[str, Any]] = []
    for row in mapping.sort_values("series_column_1_based").itertuples(index=False):
        column = int(row.series_column_1_based)
        raw = frame[column].copy()
        corrected, imputed = _isolated_imputation(raw)
        flags = np.full(len(frame), "source_value", dtype=object)
        flags[imputed] = "imputed_linear_isolated"
        missing = corrected.isna().to_numpy()
        flags[missing] = "missing_not_imputed"
        outlier = corrected.abs().gt(2.0 * float(row.capacity_mva)).fillna(False).to_numpy()
        if year == 2024:
            timestamps = pd.DatetimeIndex(frame.index)
            exact = (
                ((column == 31) & timestamps.isin(pd.to_datetime(["2024-11-23 04:00", "2024-11-29 07:00"])))
                | ((column == 32) & timestamps.isin(pd.to_datetime(["2024-12-08 15:00"])))
            )
            outlier |= exact
            for timestamp in timestamps[exact]:
                value = raw.loc[timestamp]
                issues.append(
                    {
                        "issue_id": f"TS-2024-ANOMALY-{column}-{timestamp.strftime('%Y%m%d%H%M')}",
                        "year": year,
                        "timestamp": timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                        "source_column_1_based": column,
                        "source_header_station_id": row.source_header_station_id,
                        "transformer_uid": row.transformer_uid,
                        "raw_value_mw": float(value),
                        "description": "2024 quantity outlier isolated and excluded from solving",
                        **_lineage(
                            "SRC07:Sheet3",
                            "retain raw anomaly and replace corrected value with missing",
                            "isolated_quantity_outlier",
                            source_hash,
                            scenario_id=DATASET_ID,
                        ),
                    }
                )
        corrected.iloc[np.flatnonzero(outlier)] = np.nan
        flags[outlier] = "outlier_removed_not_imputed"
        records.append(
            pd.DataFrame(
                {
                    "timestamp": frame.index.strftime("%Y-%m-%d %H:%M:%S"),
                    "source_row": source_rows.to_numpy(),
                    "year": year,
                    "source_sheet": sheet,
                    "series_column_1_based": column,
                    "source_header_station_id": row.source_header_station_id,
                    "transformer_uid": row.transformer_uid,
                    "region_id": row.region_id,
                    "voltage_kv": int(row.voltage_kv),
                    "station_id": row.station_id,
                    "unit_id": row.unit_id,
                    "net_load_mw_raw": raw.to_numpy(),
                    "net_load_mw": corrected.to_numpy(),
                    "point_quality_flag": flags,
                    "mapping_approval_status": row.approval_status,
                    "formal_use_allowed": False,
                    **_lineage(
                        f"SRC07:{sheet}",
                        "retain raw values; interpolate isolated gaps and isolate known quantity outliers",
                        ";".join(sorted(set(flags))),
                        source_hash,
                        scenario_id=DATASET_ID,
                    ),
                }
            )
        )
        if missing.any() or outlier.any() or row.approval_status != "conditional":
            issues.append(
                {
                    "issue_id": f"TS-{year}-{column:02d}-SUMMARY",
                    "year": year,
                    "timestamp": "",
                    "source_column_1_based": column,
                    "source_header_station_id": row.source_header_station_id,
                    "transformer_uid": row.transformer_uid,
                    "raw_value_mw": np.nan,
                    "description": f"mapping={row.approval_status}; missing={int(missing.sum())}; outlier={int(outlier.sum())}",
                    **_lineage(
                        f"SRC07:{sheet}",
                        "summarize annual quality flags for mapping gate",
                        f"annual_{row.approval_status}",
                        source_hash,
                        scenario_id=DATASET_ID,
                    ),
                }
            )
    return pd.concat(records, ignore_index=True), issues


def build_cross_year_timeseries_artifacts(
    source_root: Path,
    processed_root: Path,
    contract_path: Path,
    candidate_map_path: Path | None = None,
) -> dict[str, Any]:
    source_root = Path(source_root).resolve()
    processed_root = Path(processed_root).resolve()
    contract = yaml.safe_load(Path(contract_path).read_text(encoding="utf-8"))
    if str(contract["contract"]["version"]) not in {"3.1.0", "3.2.0"}:
        raise AnnualAssetScopeError(
            "cross-year artifacts require model contract 3.1.0 or 3.2.0"
        )
    candidate_path = Path(candidate_map_path or (Path(__file__).resolve().parents[1] / CANDIDATE_FILE)).resolve()
    source_hash = _sha256(source_root / HOURLY_FILE)
    mapping, approval, review_result = _candidate_records(source_root, processed_root, candidate_path, source_hash)
    headers_by_year = {year: _source_headers(source_root / HOURLY_FILE, sheet) for year, sheet in YEAR_SHEETS.items()}
    review_rows: list[dict[str, Any]] = []
    for year, sheet in YEAR_SHEETS.items():
        for row in mapping.itertuples(index=False):
            review_rows.append(
                {
                    "year": year,
                    "source_sheet": sheet,
                    "series_column_1_based": int(row.series_column_1_based),
                    "source_header_station_id": headers_by_year[year][int(row.series_column_1_based) - 1],
                    "transformer_uid": row.transformer_uid,
                    "station_id": row.station_id,
                    "unit_id": row.unit_id,
                    "voltage_kv": int(row.voltage_kv),
                    "candidate_approval_status": row.approval_status,
                    "confidence": row.confidence,
                    **_lineage(
                        f"SRC07:{sheet}+candidate_map",
                        "review cross-year source header and stable candidate target without silent header overwrite",
                        f"review_{row.approval_status}",
                        source_hash,
                        scenario_id=DATASET_ID,
                    ),
                }
            )
    review = pd.DataFrame(review_rows)
    hourly_issues: list[dict[str, Any]] = []
    processed_root.mkdir(parents=True, exist_ok=True)
    _write_frame(processed_root / "timeseries_column_map_2022_2026_review.csv", review, ["year", "series_column_1_based"])
    _write_frame(processed_root / "timeseries_mapping_approval.csv", approval, ["series_column_1_based"])
    # 保留兼容查询文件，但其状态仍是 v3 审批前的候选状态。
    _write_frame(processed_root / "timeseries_column_map_2025.csv", mapping, ["series_column_1_based"])
    for year in (2022, 2023, 2024, 2025):
        hourly, issues = _build_hourly_year(source_root, mapping, year, source_hash)
        _write_frame(processed_root / f"transformer_hourly_{year}.csv.gz", hourly, ["timestamp", "series_column_1_based"])
        hourly_issues.extend(issues)
    issue_frame = pd.DataFrame(hourly_issues)
    if not issue_frame.empty:
        _write_frame(processed_root / "timeseries_quality_issues.csv", issue_frame, ["year", "source_column_1_based", "issue_id"])
    review_result.update(
        {
            "review_rows": len(review),
            "formal_years": [2022, 2023, 2024, 2025],
            "hourly_output_rows": {year: int(pd.read_csv(processed_root / f"transformer_hourly_{year}.csv.gz").shape[0]) for year in (2022, 2023, 2024, 2025)},
            "grade_a_ready": False,
            "formal_hourly_use_allowed": False,
        }
    )
    return {"mapping": mapping, "approval": approval, "review": review, "quality_issues": issue_frame, "result": review_result}


def approve_real_timeseries_mapping(
    processed_root: Path,
    approval_path: Path,
    output_path: Path | None = None,
) -> dict[str, Any]:
    """读取项目负责人审批副本并按年度运行白名单更新正式门禁。

    2025 年 QX-00005 的正式逐时范围是 110 kV ``operating_2025`` 白名单中的
    40 台主变，而不是映射表中的全部 42 台 110 kV 主变，更不是 58 列的
    全部上下文设备。年末新增的 BDZ-00056 两台设备必须保留审批记录，但
    不能因为不具备全年时序而阻断其余 40 台设备。
    """
    processed_root = Path(processed_root).resolve()
    approval_path = Path(approval_path).resolve()
    target = Path(output_path or processed_root / "timeseries_mapping_approval.csv").resolve()
    current = pd.read_csv(processed_root / "timeseries_mapping_approval.csv")
    submitted = pd.read_csv(approval_path)
    required = {
        "series_column_1_based",
        "transformer_uid",
        "approval_status",
        "approval_authority",
        "approval_date",
        "approval_basis",
        "candidate_map_sha256",
    }
    if not required <= set(submitted.columns):
        raise AnnualAssetScopeError(f"approval file missing columns {sorted(required - set(submitted.columns))}")
    if len(submitted) != len(current) or set(submitted["series_column_1_based"]) != set(
        current["series_column_1_based"]
    ):
        raise AnnualAssetScopeError("approval must contain exactly the 58 reviewed source columns")
    if submitted["series_column_1_based"].duplicated().any():
        raise AnnualAssetScopeError("approval source columns must be unique")
    if set(submitted["approval_authority"]) != {"project_owner"}:
        raise AnnualAssetScopeError("only project_owner approval is accepted")
    if submitted["approval_status"].isna().any() or not set(submitted["approval_status"]) <= {
        "approved",
        "conditional",
        "rejected",
    }:
        raise AnnualAssetScopeError("approval contains an unsupported approval status")
    if set(submitted["candidate_map_sha256"]) != set(current["candidate_map_sha256"]):
        raise AnnualAssetScopeError("approval candidate hash does not match review")
    identity = current[["series_column_1_based", "transformer_uid"]].merge(
        submitted[["series_column_1_based", "transformer_uid"]],
        on="series_column_1_based",
        how="left",
        suffixes=("_current", "_submitted"),
        validate="one_to_one",
    )
    if identity["transformer_uid_submitted"].isna().any() or not identity["transformer_uid_current"].eq(
        identity["transformer_uid_submitted"]
    ).all():
        raise AnnualAssetScopeError("approval target transformer differs from the reviewed candidate mapping")

    whitelist_path = processed_root / "annual_asset_whitelist.csv"
    if not whitelist_path.is_file():
        raise AnnualAssetScopeError("annual_asset_whitelist.csv is required for scope-aware approval")
    whitelist = pd.read_csv(whitelist_path)
    operating_mask = (
        whitelist["year"].astype(int).eq(2025)
        & whitelist["region_id"].eq("QX-00005")
        & whitelist["voltage_kv"].astype(int).eq(110)
        & whitelist["asset_scope_id"].eq("operating_2025")
        & whitelist["in_annual_operating_whitelist"].astype(str).str.lower().eq("true")
    )
    required_uids = set(whitelist.loc[operating_mask, "transformer_uid"].astype(str))
    if len(required_uids) != 40:
        raise AnnualAssetScopeError(
            f"2025 QX-00005 110 kV operating whitelist must contain 40 transformers, got {len(required_uids)}"
        )
    current_required = set(
        current.loc[
            current["transformer_uid"].astype(str).isin(required_uids),
            "transformer_uid",
        ].astype(str)
    )
    if current_required != required_uids:
        missing = sorted(required_uids - current_required)
        extra = sorted(current_required - required_uids)
        raise AnnualAssetScopeError(f"mapping and annual operating whitelist differ; missing={missing}, extra={extra}")

    merged = current.drop(columns=["approval_status", "approval_authority", "approval_date", "approval_basis"], errors="ignore").merge(
        submitted[["series_column_1_based", "approval_status", "approval_authority", "approval_date", "approval_basis"]],
        on="series_column_1_based",
        how="left",
        validate="one_to_one",
    )
    merged["formal_use_allowed"] = (
        merged["transformer_uid"].astype(str).isin(required_uids)
        & merged["approval_status"].eq("approved")
        & merged["approval_authority"].eq("project_owner")
    )
    formal_mask = merged["formal_use_allowed"]
    approved_mask = merged["approval_status"].eq("approved")
    excluded_mask = merged["transformer_uid"].astype(str).isin(
        {
            "QX-00005|110|BDZ-00056|#1",
            "QX-00005|110|BDZ-00056|#2",
        }
    )
    merged.loc[formal_mask, "quality_flag"] = "project_owner_approved_2025_operating_110kv"
    merged.loc[approved_mask & ~formal_mask, "quality_flag"] = "project_owner_approved_context_only"
    merged.loc[excluded_mask & ~formal_mask, "quality_flag"] = "year_end_only_excluded_from_2025_operating_gate"
    merged.loc[~approved_mask & ~excluded_mask, "quality_flag"] = "project_owner_approval_incomplete"
    _write_frame(target, merged, ["series_column_1_based"])

    required_formal_rows = int(merged["transformer_uid"].astype(str).isin(required_uids).sum())
    approved_formal_rows = int(merged["formal_use_allowed"].sum())
    formal_use_allowed = bool(required_formal_rows == 40 and approved_formal_rows == 40)

    # 只有提交到 canonical processed 目录时才回写派生查询文件；测试或审查副本
    # 使用独立 output_path 时不改变其输入目录。
    canonical_target = processed_root / "timeseries_mapping_approval.csv"
    if target == canonical_target:
        status_by_column = merged.set_index("series_column_1_based")
        mapping_path = processed_root / "timeseries_column_map_2025.csv"
        if mapping_path.is_file():
            mapping = pd.read_csv(mapping_path)
            mapping["approval_status"] = mapping["series_column_1_based"].map(status_by_column["approval_status"])
            mapping["approval_authority"] = mapping["series_column_1_based"].map(status_by_column["approval_authority"])
            mapping["approval_date"] = mapping["series_column_1_based"].map(status_by_column["approval_date"])
            mapping["approval_basis"] = mapping["series_column_1_based"].map(status_by_column["approval_basis"])
            mapping["formal_use_allowed"] = mapping["series_column_1_based"].map(status_by_column["formal_use_allowed"]).fillna(False).astype(bool)
            mapping["quality_flag"] = mapping["series_column_1_based"].map(status_by_column["quality_flag"])
            _write_frame(mapping_path, mapping, ["series_column_1_based"])

        review_path = processed_root / "timeseries_column_map_2022_2026_review.csv"
        if review_path.is_file():
            review = pd.read_csv(review_path)
            review["candidate_approval_status"] = review["series_column_1_based"].map(status_by_column["approval_status"])
            review["quality_flag"] = review["series_column_1_based"].map(status_by_column["quality_flag"])
            _write_frame(review_path, review, ["year", "series_column_1_based"])

        for year in (2022, 2023, 2024, 2025):
            hourly_path = processed_root / f"transformer_hourly_{year}.csv.gz"
            if not hourly_path.is_file():
                continue
            hourly = pd.read_csv(hourly_path)
            hourly["mapping_approval_status"] = hourly["series_column_1_based"].map(status_by_column["approval_status"])
            hourly["formal_use_allowed"] = (
                hourly["year"].astype(int).eq(2025)
                & hourly["transformer_uid"].astype(str).isin(required_uids)
                & hourly["mapping_approval_status"].eq("approved")
            )
            _write_frame(hourly_path, hourly, ["timestamp", "series_column_1_based"])

        quality_path = processed_root / "timeseries_quality_issues.csv"
        if quality_path.is_file():
            quality = pd.read_csv(quality_path)
            summary_mask = quality["issue_id"].astype(str).str.endswith("-SUMMARY")
            for index in quality.index[summary_mask]:
                column = int(quality.at[index, "source_column_1_based"])
                status = str(status_by_column.loc[column, "approval_status"])
                parts = str(quality.at[index, "description"]).split(";")
                if parts and parts[0].startswith("mapping="):
                    parts[0] = f"mapping={status}"
                quality.at[index, "description"] = ";".join(parts)
                quality.at[index, "quality_flag"] = str(status_by_column.loc[column, "quality_flag"])
            _write_frame(quality_path, quality, ["year", "source_column_1_based", "issue_id"])

        quality_ledger_path = processed_root / "data_quality_issues.csv"
        if quality_ledger_path.is_file():
            quality_ledger = pd.read_csv(quality_ledger_path)
            summary_mask = quality_ledger["issue_id"].astype(str).str.endswith("-SUMMARY")
            for index in quality_ledger.index[summary_mask]:
                column = int(quality_ledger.at[index, "source_column_1_based"])
                status = str(status_by_column.loc[column, "approval_status"])
                parts = str(quality_ledger.at[index, "description"]).split(";")
                if parts and parts[0].startswith("mapping="):
                    parts[0] = f"mapping={status}"
                quality_ledger.at[index, "description"] = ";".join(parts)
                quality_ledger.at[index, "quality_flag"] = str(status_by_column.loc[column, "quality_flag"])
            candidate_issue = quality_ledger["issue_id"].astype(str).eq("DQ-TIMESERIES-CANDIDATE")
            quality_ledger.loc[candidate_issue, "severity"] = "warning_for_scope_exceptions"
            quality_ledger.loc[candidate_issue, "description"] = (
                "2025 time-series mapping is approved for 40 operating 110 kV transformers; 16 35 kV rows are "
                "approved as context only and the two BDZ-00056 year-end-only rows are excluded from the operating gate."
            )
            quality_ledger.loc[candidate_issue, "quality_flag"] = "project_owner_approved_with_scope_exceptions"
            gate_issue = quality_ledger["issue_id"].astype(str).eq("DQ-V3-MAPPING-APPROVAL")
            quality_ledger.loc[gate_issue, "severity"] = "warning_for_scope_exceptions"
            quality_ledger.loc[gate_issue, "description"] = (
                "56 reviewed source columns are approved by project_owner; the two BDZ-00056 year-end-only "
                "columns remain excluded from the 2025 QX-00005 110 kV operating whitelist, while historical "
                "completeness and 2024 anomaly isolation remain disclosed."
            )
            quality_ledger.loc[gate_issue, "quality_flag"] = "project_owner_approved_with_scope_exceptions"
            _write_frame(quality_ledger_path, quality_ledger, ["issue_id"])

        manifest_path = processed_root / "manifest.json"
        if manifest_path.is_file():
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            gate = manifest.setdefault("timeseries_gate", {})
            gate.update(
                {
                    "grade_a_ready": formal_use_allowed,
                    "formal_hourly_use_allowed": formal_use_allowed,
                    "formal_scope_id": "2025_QX-00005_110kv_operating",
                    "formal_scope_required_transformers": required_formal_rows,
                    "formal_scope_approved_transformers": approved_formal_rows,
                    "approved_mapping_rows": int(approved_mask.sum()),
                    "rejected_mapping_rows": int((~approved_mask).sum()),
                    "year_end_only_excluded_mapping_rows": int(excluded_mask.sum()),
                }
            )
            manifest.setdefault("timeseries_review", {}).update(
                {
                    "grade_a_ready": formal_use_allowed,
                    "formal_hourly_use_allowed": formal_use_allowed,
                    "approved_rows": int(approved_mask.sum()),
                    "approved_columns": int(approved_mask.sum()),
                    "conditional_columns": int(merged["approval_status"].eq("conditional").sum()),
                    "rejected_columns": int(merged["approval_status"].eq("rejected").sum()),
                    "required_formal_rows": required_formal_rows,
                    "approved_formal_rows": approved_formal_rows,
                }
            )
            manifest["output_files"] = {
                path.name: {"sha256": _sha256(path), "bytes": path.stat().st_size}
                for path in sorted(processed_root.iterdir(), key=lambda item: item.name)
                if path.is_file() and path.name != "manifest.json"
            }
            manifest_path.write_text(
                json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
                encoding="utf-8",
            )

    return {
        "approval_file": str(target),
        "approved_rows": int(approved_mask.sum()),
        "total_rows": int(len(merged)),
        "grade_a_ready": formal_use_allowed,
        "required_formal_rows": required_formal_rows,
        "approved_formal_rows": approved_formal_rows,
        "excluded_year_end_rows": int(excluded_mask.sum()),
        "approval": merged,
        "formal_use_allowed": formal_use_allowed,
    }
