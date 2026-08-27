from __future__ import annotations

from pathlib import Path
import sys

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.optimize import linear_sum_assignment


root = Path(sys.argv[1])

workbook = load_workbook(root / "邳州主变负载率.xlsx", read_only=True, data_only=True)
try:
    sheet = workbook["Sheet4"]
    timestamps = []
    values = []
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=59, values_only=True):
        timestamps.append(pd.to_datetime(row[0]))
        values.append([pd.to_numeric(value, errors="coerce") for value in row[1:]])
finally:
    workbook.close()
frame = pd.DataFrame(values, index=pd.DatetimeIndex(timestamps))

workbook = load_workbook(root / "光伏装机.xlsx", read_only=True, data_only=True)
try:
    sheet = workbook["主变（跨区合并）1"]
    pv_rows = [
        row
        for row in sheet.iter_rows(min_row=2, max_row=521, values_only=True)
        if row[0] is not None
        and row[1] == "QX-00005"
        and str(row[3]).replace("kV", "") in {"35", "110"}
    ]
finally:
    workbook.close()

cost = np.zeros((58, 58))
observed = np.zeros((58, 58))
observed_time: list[list[pd.Timestamp]] = [[pd.NaT for _ in pv_rows] for _ in range(58)]
for column in range(58):
    series = frame[column]
    for row_index, row in enumerate(pv_rows):
        timestamp = pd.to_datetime(row[4])
        target = float(row[5])
        capacity = float(row[6])
        candidate_times = pd.DatetimeIndex(sorted({timestamp.floor("h"), timestamp.ceil("h")}))
        candidates = series.reindex(candidate_times).dropna()
        if candidates.empty:
            cost[column, row_index] = 1000
            observed[column, row_index] = np.nan
            continue
        best_time = (candidates - target).abs().idxmin()
        value = float(candidates.loc[best_time])
        observed[column, row_index] = value
        observed_time[column][row_index] = best_time
        cost[column, row_index] = abs(value - target) / max(capacity, 10.0)

assigned_columns, assigned_rows = linear_sum_assignment(cost)
results = []
for column, row_index in zip(assigned_columns, assigned_rows, strict=True):
    row = pv_rows[row_index]
    results.append(
        {
            "column": int(column + 1),
            "station": row[2],
            "voltage": row[3],
            "rated_mva": float(row[6]),
            "pv_row": int(row[0]),
            "target_time": str(row[4]),
            "target_net_mw": float(row[5]),
            "observed_time": str(observed_time[column][row_index]),
            "observed_mw": float(observed[column, row_index]),
            "absolute_error_mw": float(abs(observed[column, row_index] - float(row[5]))),
            "normalized_error": float(cost[column, row_index]),
        }
    )

print("summary", {
    "rows": len(results),
    "error_le_0_05_mw": sum(item["absolute_error_mw"] <= 0.05 for item in results),
    "error_le_0_5_mw": sum(item["absolute_error_mw"] <= 0.5 for item in results),
    "error_le_2_mw": sum(item["absolute_error_mw"] <= 2 for item in results),
    "median_abs_error_mw": float(np.median([item["absolute_error_mw"] for item in results])),
    "max_abs_error_mw": max(item["absolute_error_mw"] for item in results),
})
for item in sorted(results, key=lambda value: value["column"]):
    print(item)
