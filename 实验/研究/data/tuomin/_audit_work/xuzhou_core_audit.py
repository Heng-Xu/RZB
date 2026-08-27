from __future__ import annotations

import json
import math
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import networkx as nx
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.optimize import linear_sum_assignment


ROOT = Path(sys.argv[1])
COUNTY = "QX-00005"


def clean(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def voltage(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value).lower().replace("kv", "").replace("千伏", "").strip()
    try:
        return int(float(text))
    except ValueError:
        return None


def number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def unit(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip().replace("＃", "#")
    if text and not text.startswith("#"):
        text = "#" + text
    return text or None


def rows(
    workbook_path: Path,
    sheet_name: str,
    min_row: int,
    max_row: int | None = None,
    max_col: int | None = None,
) -> list[tuple[Any, ...]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        return [
            tuple(clean(value) for value in row)
            for row in sheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                max_col=max_col,
                values_only=True,
            )
        ]
    finally:
        workbook.close()


def rounded(value: Any, digits: int = 4) -> Any:
    if isinstance(value, (float, np.floating)):
        return round(float(value), digits)
    return value


def emit(title: str, payload: Any) -> None:
    print(f"\n===== {title} =====")
    print(json.dumps(payload, ensure_ascii=False, indent=2, default=str))


load_path = ROOT / "2025设备负载统计表.xlsx"
equipment_path = ROOT / "110（35）kv设备明细.xlsx"
pv_path = ROOT / "光伏装机.xlsx"
profile_path = ROOT / "光伏8760小时数据.xlsx"
feeder_path = ROOT / "邳州10kV线路基础数据表.xlsx"
hourly_path = ROOT / "邳州主变负载率.xlsx"
history_path = ROOT / "近5年容载比.xlsx"


# SRC01: station-level annual statistics. Sequence number is the only reliable used-range guard.
station_all = [row for row in rows(load_path, "变电站1", 2, 1000, 23) if number(row[0]) is not None]
station_scope = [row for row in station_all if voltage(row[1]) == 110 and row[2] == COUNTY]
station_ids = {row[3] for row in station_scope}
station_capacity = {row[3]: number(row[4]) or 0.0 for row in station_scope}
station_max = {row[3]: number(row[12]) for row in station_scope}
station_min = {row[3]: number(row[15]) for row in station_scope}
static_bidirectional_clr = {
    station_id: station_capacity[station_id]
    / max(abs(station_max[station_id] or 0.0), abs(station_min[station_id] or 0.0))
    if max(abs(station_max[station_id] or 0.0), abs(station_min[station_id] or 0.0)) > 0
    else None
    for station_id in station_ids
}
emit(
    "SRC01_STATIONS",
    {
        "valid_rows_all_voltage_county": len(station_all),
        "rows_by_voltage": dict(sorted(Counter(voltage(row[1]) for row in station_all).items(), key=str)),
        "scope_rows": len(station_scope),
        "scope_unique_station_ids": len(station_ids),
        "duplicate_station_ids": sorted(
            key for key, count in Counter(row[3] for row in station_scope).items() if count > 1
        ),
        "capacity_sum_mva": rounded(sum(station_capacity.values())),
        "missing_capacity": sorted(row[3] for row in station_scope if number(row[4]) is None),
        "missing_annual_max": sorted(row[3] for row in station_scope if number(row[12]) is None),
        "missing_annual_min": sorted(row[3] for row in station_scope if number(row[15]) is None),
        "reverse_flow_station_count": sum((number(row[15]) or 0) < 0 for row in station_scope),
        "zero_or_missing_extrema_station_ids": sorted(
            row[3]
            for row in station_scope
            if not number(row[12]) and not number(row[15])
        ),
        "static_station_clr_range": [
            rounded(min(value for value in static_bidirectional_clr.values() if value is not None)),
            rounded(max(value for value in static_bidirectional_clr.values() if value is not None)),
        ],
        "highest_static_clr": [
            {"station_id": key, "clr": rounded(value)}
            for key, value in sorted(
                ((key, value) for key, value in static_bidirectional_clr.items() if value is not None),
                key=lambda item: item[1],
                reverse=True,
            )[:5]
        ],
    },
)


# SRC02: transformer-level annual statistics.
transformer_all = [row for row in rows(load_path, "主变1", 3, 522, 30) if number(row[0]) is not None]
transformer_county = [row for row in transformer_all if row[2] == COUNTY and voltage(row[1]) in {35, 110}]
transformer_scope = [row for row in transformer_county if voltage(row[1]) == 110]
transformer35_scope = [row for row in transformer_county if voltage(row[1]) == 35]
transformer_keys = {(row[3], unit(row[4])) for row in transformer_scope}
transformer_capacity_by_station: dict[str, float] = defaultdict(float)
for row in transformer_scope:
    transformer_capacity_by_station[row[3]] += number(row[5]) or 0.0
capacity_diffs = {
    station_id: rounded(transformer_capacity_by_station.get(station_id, 0.0) - capacity)
    for station_id, capacity in station_capacity.items()
    if abs(transformer_capacity_by_station.get(station_id, 0.0) - capacity) > 1e-6
}
emit(
    "SRC02_TRANSFORMERS",
    {
        "all_transformer_rows": len(transformer_all),
        "scope_110_rows": len(transformer_scope),
        "scope_35_rows": len(transformer35_scope),
        "scope_110_unique_keys": len(transformer_keys),
        "scope_110_station_count": len({row[3] for row in transformer_scope}),
        "units_per_station": dict(sorted(Counter(row[3] for row in transformer_scope).items())),
        "capacity_sum_mva": rounded(sum(number(row[5]) or 0.0 for row in transformer_scope)),
        "station_capacity_differences_mva": capacity_diffs,
        "missing_annual_max_keys": [
            f"{row[3]}/{unit(row[4])}" for row in transformer_scope if number(row[14]) is None
        ],
        "missing_annual_min_keys": [
            f"{row[3]}/{unit(row[4])}" for row in transformer_scope if number(row[17]) is None
        ],
        "negative_min_transformer_count": sum((number(row[17]) or 0) < 0 for row in transformer_scope),
        "heavy_transformer_count": sum(clean(row[20]) is not None for row in transformer_scope),
    },
)


# SRC03: equipment and expansion constraints.
equipment_all = [row for row in rows(equipment_path, "110千伏变电站1", 4, 320, 47) if number(row[0]) is not None]
equipment_scope = [row for row in equipment_all if row[1] == COUNTY and voltage(row[5]) == 110]
equipment_keys = {(row[3], unit(row[14])) for row in equipment_scope}
equipment_capacity = {(row[3], unit(row[14])): number(row[15]) for row in equipment_scope}
transformer_capacity = {(row[3], unit(row[4])): number(row[5]) for row in transformer_scope}
key_missing_in_equipment = sorted(f"{a}/{b}" for a, b in transformer_keys - equipment_keys)
key_extra_in_equipment = sorted(f"{a}/{b}" for a, b in equipment_keys - transformer_keys)
equipment_capacity_conflicts = {
    f"{key[0]}/{key[1]}": [transformer_capacity.get(key), equipment_capacity.get(key)]
    for key in transformer_keys & equipment_keys
    if transformer_capacity.get(key) != equipment_capacity.get(key)
}
station_equipment: dict[str, list[tuple[Any, ...]]] = defaultdict(list)
for row in equipment_scope:
    station_equipment[row[3]].append(row)


def first_number(records: Iterable[tuple[Any, ...]], index: int) -> float | None:
    return next((number(record[index]) for record in records if number(record[index]) is not None), None)


emit(
    "SRC03_EQUIPMENT",
    {
        "scope_rows": len(equipment_scope),
        "scope_stations": len(station_equipment),
        "key_missing_in_equipment": key_missing_in_equipment,
        "key_extra_in_equipment": key_extra_in_equipment,
        "capacity_conflicts": equipment_capacity_conflicts,
        "expandable_station_count": sum(
            (first_number(records, 10) or 0) > 0 for records in station_equipment.values()
        ),
        "expandable_capacity_sum_mva": rounded(
            sum(first_number(records, 11) or 0 for records in station_equipment.values())
        ),
        "remaining_10kv_bays_sum": rounded(
            sum(first_number(records, 13) or 0 for records in station_equipment.values())
        ),
        "stations_with_no_remaining_10kv_bay": sorted(
            station_id
            for station_id, records in station_equipment.items()
            if (first_number(records, 13) or 0) <= 0
        ),
        "n_minus_1_not_yes": sorted(
            {row[3] for row in equipment_scope if clean(row[31]) not in {"是", None}}
        ),
    },
)


# SRC04: 110 kV capacity network.
line_all = [row for row in rows(equipment_path, "110千伏线路1", 3, 366, 20) if number(row[0]) is not None]
line_touching = [row for row in line_all if row[12] in station_ids or row[13] in station_ids]
line_county = [row for row in line_all if row[1] == COUNTY]
line_ids = [row[2] for row in line_touching]
boundary_nodes = sorted(
    {node for row in line_touching for node in (row[12], row[13]) if node not in station_ids and node is not None}
)
graph = nx.Graph()
graph.add_nodes_from(station_ids)
for row in line_touching:
    if row[12] and row[13]:
        graph.add_edge(row[12], row[13], line_id=row[2])
research_subgraph = graph.subgraph(station_ids)
emit(
    "SRC04_NETWORK",
    {
        "all_110kv_lines": len(line_all),
        "county_tagged_lines": len(line_county),
        "lines_touching_research_station": len(line_touching),
        "unique_touching_line_ids": len(set(line_ids)),
        "duplicate_touching_line_ids": sorted(
            line_id for line_id, count in Counter(line_ids).items() if count > 1
        ),
        "boundary_node_count": len(boundary_nodes),
        "boundary_nodes": boundary_nodes,
        "missing_or_zero_current_limit_lines": sorted(
            row[2] for row in line_touching if not number(row[11])
        ),
        "missing_length_lines": sorted(row[2] for row in line_touching if number(row[6]) is None),
        "max_load_rate_over_100_lines": sorted(
            row[2] for row in line_touching if (number(row[15]) or 0) > 100
        ),
        "research_only_connected_components": [
            sorted(component) for component in nx.connected_components(research_subgraph)
        ],
        "research_station_degree_in_full_touching_graph": dict(
            sorted((station_id, graph.degree(station_id)) for station_id in station_ids)
        ),
    },
)


# SRC05: station-aggregated PV snapshot.
pv_all = [row for row in rows(pv_path, "主变（跨区合并）1", 2, 521, 10) if number(row[0]) is not None]
pv_scope = [row for row in pv_all if row[1] == COUNTY and voltage(row[3]) == 110]
pv_by_station: dict[str, dict[str, float]] = defaultdict(lambda: {"online": 0.0, "pipeline": 0.0, "rated": 0.0})
for row in pv_scope:
    pv_by_station[row[2]]["rated"] += number(row[6]) or 0.0
    pv_by_station[row[2]]["online"] += number(row[7]) or 0.0
    pv_by_station[row[2]]["pipeline"] += number(row[8]) or 0.0
pv_times = pd.to_datetime([row[4] for row in pv_scope], errors="coerce")
emit(
    "SRC05_PV_CAPACITY",
    {
        "scope_rows": len(pv_scope),
        "scope_station_count": len(pv_by_station),
        "rows_per_station": dict(sorted(Counter(row[2] for row in pv_scope).items())),
        "missing_station_ids": sorted(station_ids - set(pv_by_station)),
        "extra_station_ids": sorted(set(pv_by_station) - station_ids),
        "online_total_mw": rounded(sum(item["online"] for item in pv_by_station.values())),
        "pipeline_total_mw": rounded(sum(item["pipeline"] for item in pv_by_station.values())),
        "online_plus_pipeline_total_mw": rounded(
            sum(item["online"] + item["pipeline"] for item in pv_by_station.values())
        ),
        "rated_capacity_total_mva": rounded(sum(item["rated"] for item in pv_by_station.values())),
        "rated_capacity_difference_vs_src01_mva": rounded(
            sum(item["rated"] for item in pv_by_station.values()) - sum(station_capacity.values())
        ),
        "snapshot_time_min": pv_times.min(),
        "snapshot_time_max": pv_times.max(),
        "negative_capacity_rows": [
            int(row[0])
            for row in pv_scope
            if (number(row[7]) or 0) < 0 or (number(row[8]) or 0) < 0
        ],
        "station_totals": {
            key: {field: rounded(value) for field, value in values.items()}
            for key, values in sorted(pv_by_station.items())
        },
    },
)


# SRC06: feeder assets and consistency checks.
feeder_all = [row for row in rows(feeder_path, "邳州10千伏线路基础数据", 2, 404, 19) if number(row[0]) is not None]
feeder_scope = [row for row in feeder_all if row[12] in station_ids]
reverse_yes = [row for row in feeder_scope if row[11] == "是"]
negative_active = [row for row in feeder_scope if (number(row[9]) or 0) < 0]
length_mismatch = [
    row[1]
    for row in feeder_scope
    if number(row[13]) is not None
    and number(row[14]) is not None
    and number(row[15]) is not None
    and abs((number(row[14]) or 0) + (number(row[15]) or 0) - (number(row[13]) or 0)) > 0.01
]
rate_mismatch = [
    row[1]
    for row in feeder_scope
    if number(row[7])
    and number(row[8]) is not None
    and number(row[10]) is not None
    and abs((number(row[8]) or 0) / (number(row[7]) or 1) * 100 - (number(row[10]) or 0)) > 0.05
]
emit(
    "SRC06_FEEDERS",
    {
        "all_rows": len(feeder_all),
        "unique_feeder_ids": len({row[1] for row in feeder_all}),
        "duplicate_feeder_ids": sorted(
            feeder_id for feeder_id, count in Counter(row[1] for row in feeder_all).items() if count > 1
        ),
        "research_station_feeder_rows": len(feeder_scope),
        "research_stations_with_feeders": len({row[12] for row in feeder_scope}),
        "research_stations_without_feeders": sorted(station_ids - {row[12] for row in feeder_scope}),
        "feeders_per_station": dict(sorted(Counter(row[12] for row in feeder_scope).items())),
        "negative_max_active_count": len(negative_active),
        "reverse_flag_yes_count": len(reverse_yes),
        "negative_without_reverse_flag": sorted(row[1] for row in negative_active if row[11] != "是"),
        "reverse_flag_without_negative_active": sorted(row[1] for row in reverse_yes if (number(row[9]) or 0) >= 0),
        "missing_max_current": sum(number(row[8]) is None for row in feeder_scope),
        "missing_max_active": sum(number(row[9]) is None for row in feeder_scope),
        "length_component_mismatch_count": len(length_mismatch),
        "length_component_mismatch_examples": length_mismatch[:20],
        "load_rate_mismatch_count": len(rate_mismatch),
        "load_rate_mismatch_examples": rate_mismatch[:20],
        "connected_transformer_capacity_kva_sum": rounded(
            sum(number(row[16]) or 0 for row in feeder_scope)
        ),
    },
)


# SRC14: PV profile integrity and physical caveat checks.
profile_rows = rows(profile_path, "光伏8760", 2, 8761, 6)
profile_time = pd.to_datetime([row[1] for row in profile_rows], errors="coerce")
profile_phi = pd.to_numeric(pd.Series([row[4] for row in profile_rows]), errors="coerce")
profile_scaled = pd.to_numeric(pd.Series([row[5] for row in profile_rows]), errors="coerce")
profile_diff = pd.Series(profile_time).diff().dropna()
night_mask = pd.Series(profile_time).dt.hour.isin(list(range(0, 6)) + list(range(20, 24)))
emit(
    "SRC14_PV_PROFILE",
    {
        "row_count": len(profile_rows),
        "timestamp_unique_count": int(pd.Series(profile_time).nunique()),
        "timestamp_missing_count": int(pd.Series(profile_time).isna().sum()),
        "start": profile_time.min(),
        "end": profile_time.max(),
        "non_hourly_gap_count": int((profile_diff != pd.Timedelta(hours=1)).sum()),
        "phi_missing_count": int(profile_phi.isna().sum()),
        "phi_min": rounded(profile_phi.min(), 9),
        "phi_max": rounded(profile_phi.max(), 9),
        "phi_mean": rounded(profile_phi.mean(), 9),
        "equivalent_full_load_hours": rounded(profile_phi.sum(), 2),
        "out_of_range_count": int(((profile_phi < 0) | (profile_phi > 1)).sum()),
        "scaled_3408_max_abs_error_mw": rounded((profile_scaled - profile_phi * 3408).abs().max(), 10),
        "night_hours": int(night_mask.sum()),
        "night_positive_count": int((profile_phi[night_mask] > 0).sum()),
        "night_phi_max": rounded(profile_phi[night_mask].max(), 9),
        "exact_zero_count": int((profile_phi == 0).sum()),
    },
)


# SRC07: workbook-wide time and missingness audit; keep 2025 for mapping diagnostics.
hourly_book = load_workbook(hourly_path, read_only=True, data_only=True)
hourly_summaries: dict[str, Any] = {}
year2025: pd.DataFrame | None = None
header_units: list[str | None] | None = None
try:
    for sheet in hourly_book.worksheets:
        row1 = [clean(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        row2 = [clean(value) for value in next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))]
        current_units = [unit(value) for value in row2[1:59]]
        if header_units is None:
            header_units = current_units
        records: list[tuple[pd.Timestamp, list[float]]] = []
        parse_error_cells = 0
        for raw in sheet.iter_rows(min_row=3, max_row=sheet.max_row, max_col=59, values_only=True):
            timestamp = pd.to_datetime(clean(raw[0]), errors="coerce")
            values: list[float] = []
            for value in raw[1:59]:
                if value is None or str(value).strip() == "":
                    values.append(float("nan"))
                    continue
                try:
                    values.append(float(value))
                except (TypeError, ValueError):
                    values.append(float("nan"))
                    parse_error_cells += 1
            records.append((timestamp, values))
        timestamps = pd.DatetimeIndex([record[0] for record in records])
        frame = pd.DataFrame([record[1] for record in records], index=timestamps)
        diffs = pd.Series(timestamps).diff().dropna()
        all_missing_rows = frame.isna().all(axis=1)
        yearly_counts = Counter(int(ts.year) for ts in timestamps if not pd.isna(ts))
        hourly_summaries[sheet.title] = {
            "row_count": len(frame),
            "data_columns": frame.shape[1],
            "start": timestamps.min(),
            "end": timestamps.max(),
            "timestamp_unique": int(timestamps.nunique()),
            "timestamp_missing": int(timestamps.isna().sum()),
            "duplicate_timestamp_count": int(timestamps.duplicated().sum()),
            "non_hourly_gap_count": int((diffs != pd.Timedelta(hours=1)).sum()),
            "year_counts": dict(sorted(yearly_counts.items())),
            "all_series_missing_row_count": int(all_missing_rows.sum()),
            "all_series_missing_timestamps": [str(ts) for ts in frame.index[all_missing_rows][:30]],
            "parse_error_cells": parse_error_cells,
            "active_column_count": int((frame.notna().sum(axis=0) > 0).sum()),
            "fully_missing_columns_1_based": [
                int(index + 1) for index, count in frame.notna().sum(axis=0).items() if count == 0
            ],
            "partially_missing_columns": {
                str(int(index + 1)): int(count)
                for index, count in frame.isna().sum(axis=0).items()
                if 0 < count < len(frame)
            },
            "min_value": rounded(frame.min().min()),
            "max_value": rounded(frame.max().max()),
            "negative_value_count": int((frame < 0).sum().sum()),
            "row1_unique_station_labels": sorted({str(value) for value in row1[1:59] if value is not None}),
        }
        if 2025 in yearly_counts:
            year2025 = frame[frame.index.year == 2025].copy()
finally:
    hourly_book.close()
emit("SRC07_HOURLY", hourly_summaries)


# Diagnostic-only candidate mapping. This never becomes an approved mapping automatically.
mapping_summary: dict[str, Any] = {}
if year2025 is not None and header_units is not None:
    candidates = transformer_scope + transformer35_scope
    candidate_labels = [f"{row[3]}/{unit(row[4])}/{voltage(row[1])}kV" for row in candidates]
    active_columns = [column for column in year2025.columns if year2025[column].notna().any()]
    costs = np.full((len(active_columns), len(candidates)), 1000.0)
    component_errors: dict[tuple[int, int], tuple[float, float, float, float]] = {}
    for active_index, column in enumerate(active_columns):
        series = year2025[column].dropna()
        observed_max = float(series.max())
        observed_min = float(series.min())
        for candidate_index, row in enumerate(candidates):
            expected_max = number(row[14])
            expected_min = number(row[17])
            if expected_max is None or expected_min is None:
                continue
            scale = max(abs(expected_max), abs(expected_min), number(row[5]) or 0.0, 10.0)
            max_error = abs(observed_max - expected_max) / scale
            min_error = abs(observed_min - expected_min) / scale
            at_max_error = 0.0
            at_min_error = 0.0
            for target_time, target_value, kind in (
                (row[16], expected_max, "max"),
                (row[18], expected_min, "min"),
            ):
                parsed_time = pd.to_datetime(target_time, errors="coerce")
                if pd.isna(parsed_time):
                    continue
                nearest = series.index.get_indexer([parsed_time], method="nearest", tolerance=pd.Timedelta(minutes=40))[0]
                if nearest >= 0:
                    point_error = abs(float(series.iloc[nearest]) - target_value) / scale
                    if kind == "max":
                        at_max_error = point_error
                    else:
                        at_min_error = point_error
            unit_penalty = 0.0
            header_unit = header_units[column]
            candidate_unit = unit(row[4])
            if header_unit is not None and candidate_unit is not None and header_unit != candidate_unit:
                unit_penalty = 0.75
            cost = max_error + min_error + 0.5 * (at_max_error + at_min_error) + unit_penalty
            costs[active_index, candidate_index] = cost
            component_errors[(active_index, candidate_index)] = (
                max_error,
                min_error,
                at_max_error,
                at_min_error,
            )
    assigned_rows, assigned_cols = linear_sum_assignment(costs)
    assignments = []
    for active_index, candidate_index in zip(assigned_rows, assigned_cols, strict=True):
        column = active_columns[active_index]
        sorted_costs = np.sort(costs[active_index])
        margin = float(sorted_costs[1] - sorted_costs[0]) if len(sorted_costs) > 1 else float("nan")
        errors = component_errors.get((active_index, candidate_index), (math.nan,) * 4)
        assignments.append(
            {
                "series_column_1_based": int(column + 1),
                "header_unit": header_units[column],
                "candidate": candidate_labels[candidate_index],
                "score": rounded(costs[active_index, candidate_index], 5),
                "independent_best_margin": rounded(margin, 5),
                "max_error_fraction": rounded(errors[0], 5),
                "min_error_fraction": rounded(errors[1], 5),
            }
        )
    mapping_summary = {
        "candidate_transformers": len(candidates),
        "active_2025_series": len(active_columns),
        "unassigned_candidates": sorted(set(candidate_labels) - {item["candidate"] for item in assignments}),
        "assigned_score_below_0_10": sum(item["score"] < 0.10 for item in assignments),
        "assigned_score_below_0_25": sum(item["score"] < 0.25 for item in assignments),
        "assigned_score_below_0_50": sum(item["score"] < 0.50 for item in assignments),
        "unit_header_present_columns": sum(value is not None for value in header_units),
        "assignments": sorted(assignments, key=lambda item: item["series_column_1_based"]),
        "warning": "The assignment is diagnostic evidence only; source approval is still required.",
    }
emit("SRC07_MAPPING_DIAGNOSTIC", mapping_summary)


# SRC08: historical CLR reconciliation for the research county.
history_rows = rows(history_path, "Sheet1", 4, 48, 17)
current_region: str | None = None
history_scope: dict[str, dict[int, dict[str, float | None]]] = defaultdict(dict)
for row in history_rows:
    if row[0] is not None:
        current_region = str(row[0])
    level = voltage(row[1])
    if current_region is None or level not in {35, 110}:
        continue
    years = {}
    for offset, year in zip(range(2, 17, 3), range(2021, 2026), strict=True):
        years[year] = {
            "capacity_10kva": number(row[offset]),
            "load_10kw": number(row[offset + 1]),
            "clr_reported": number(row[offset + 2]),
            "clr_recomputed": rounded((number(row[offset]) or 0) / (number(row[offset + 1]) or 1), 4),
        }
    history_scope[f"{current_region}/{level}kV"] = years
history_qx_2025 = history_scope[f"{COUNTY}/110kV"][2025]
emit(
    "SRC08_HISTORY_CLR",
    {
        "series_count": len(history_scope),
        "research_county_110kv": history_scope[f"{COUNTY}/110kV"],
        "2025_capacity_converted_mva": rounded((history_qx_2025["capacity_10kva"] or 0) * 10),
        "2025_capacity_difference_vs_src01_mva": rounded(
            (history_qx_2025["capacity_10kva"] or 0) * 10 - sum(station_capacity.values())
        ),
        "reported_recomputed_max_abs_difference": rounded(
            max(
                abs((entry["clr_reported"] or 0) - (entry["clr_recomputed"] or 0))
                for years in history_scope.values()
                for entry in years.values()
            ),
            4,
        ),
    },
)


emit(
    "KEY_STATION_DETAIL",
    [
        {
            "station_id": station_id,
            "capacity_mva": rounded(station_capacity[station_id]),
            "annual_max_mw": rounded(station_max[station_id]),
            "annual_min_mw": rounded(station_min[station_id]),
            "static_bidirectional_clr": rounded(static_bidirectional_clr[station_id]),
            "pv_online_mw": rounded(pv_by_station.get(station_id, {}).get("online", 0.0)),
            "pv_pipeline_mw": rounded(pv_by_station.get(station_id, {}).get("pipeline", 0.0)),
            "feeder_count": sum(row[12] == station_id for row in feeder_scope),
        }
        for station_id in sorted(station_ids)
    ],
)
