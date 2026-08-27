from __future__ import annotations

from pathlib import Path
import sys

import numpy as np
from openpyxl import load_workbook


path = Path(sys.argv[1]) / "逐月分县分布式光伏.xlsx"
workbook = load_workbook(path, read_only=True, data_only=True)
try:
    sheet = workbook["Sheet1"]
    blocks = {
        2023: (3, 11, 12),
        2024: (15, 23, 12),
        2025: (28, 36, 12),
        2026: (41, 49, 6),
    }
    data = {}
    for year, (start, end, months) in blocks.items():
        records = {}
        for row in sheet.iter_rows(min_row=start, max_row=end, min_col=1, max_col=months + 1, values_only=True):
            records[row[0]] = [float(value) for value in row[1:]]
        data[year] = records
finally:
    workbook.close()

for year, records in data.items():
    county_keys = sorted(key for key in records if key != "all")
    total = np.array(records["all"])
    county_sum = np.sum([records[key] for key in county_keys], axis=0)
    qx = np.array(records["QX-00005"])
    print(
        year,
        {
            "months": len(qx),
            "qx_values_raw": qx.tolist(),
            "qx_start": float(qx[0]),
            "qx_end": float(qx[-1]),
            "qx_change": float(qx[-1] - qx[0]),
            "qx_monthly_decrease_count": int((np.diff(qx) < 0).sum()),
            "all_vs_counties_max_abs_diff": float(np.max(np.abs(total - county_sum))),
        },
    )

year_ends = {year: records["QX-00005"][-1] for year, records in data.items()}
print("year_ends", year_ends)
print(
    "growth",
    {
        "2023_to_2024": year_ends[2024] / year_ends[2023] - 1,
        "2024_to_2025": year_ends[2025] / year_ends[2024] - 1,
        "2025_dec_to_2026_jun": year_ends[2026] / year_ends[2025] - 1,
    },
)
