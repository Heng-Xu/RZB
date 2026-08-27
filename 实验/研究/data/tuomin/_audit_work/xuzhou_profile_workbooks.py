from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def display(value: object, limit: int = 120) -> object:
    if value is None or isinstance(value, (int, float, bool)):
        return value
    text = str(value).replace("\n", "\\n").replace("\r", "\\r")
    return text if len(text) <= limit else text[: limit - 3] + "..."


def main(root: Path, selected: list[str]) -> None:
    paths = [root / name for name in selected] if selected else sorted(root.glob("*.xlsx"))
    for path in sorted(paths, key=lambda item: item.name):
        print(f"\n===== {path.name} ({path.stat().st_size} bytes) =====")
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                print(
                    json.dumps(
                        {
                            "sheet": sheet.title,
                            "max_row": sheet.max_row,
                            "max_column": sheet.max_column,
                            "dimension": sheet.calculate_dimension(),
                            "sheet_state": sheet.sheet_state,
                        },
                        ensure_ascii=False,
                    )
                )
                for row_number, row in enumerate(
                    sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 8), values_only=True),
                    start=1,
                ):
                    print(
                        f"row {row_number}: "
                        + json.dumps([display(value) for value in row], ensure_ascii=False)
                    )
        finally:
            workbook.close()


if __name__ == "__main__":
    main(Path(sys.argv[1]), sys.argv[2:])
