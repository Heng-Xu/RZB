from pathlib import Path
import sys

import pandas as pd
from openpyxl import load_workbook


path = Path(sys.argv[1]) / "邳州主变负载率.xlsx"
workbook = load_workbook(path, read_only=True, data_only=True)
try:
    sheet = workbook["Sheet3"]
    timestamps = []
    values = []
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=59, values_only=True):
        timestamps.append(pd.to_datetime(row[0]))
        values.append([pd.to_numeric(value, errors="coerce") for value in row[1:]])
finally:
    workbook.close()

frame = pd.DataFrame(values, index=pd.DatetimeIndex(timestamps))
print("extreme_columns")
for column in frame.columns:
    series = frame[column].dropna()
    if not series.empty and max(abs(float(series.min())), abs(float(series.max()))) > 100:
        extreme = series[series.abs() > 100]
        print(
            {
                "column": int(column + 1),
                "count_abs_gt_100": len(extreme),
                "min": float(series.min()),
                "max": float(series.max()),
                "first_extremes": {str(key): value for key, value in extreme.head(10).items()},
            }
        )
for column, timestamp in ((30, "2024-11-23 04:00:00"), (30, "2024-11-29 07:00:00"), (31, "2024-12-08 15:00:00")):
    center = pd.Timestamp(timestamp)
    print(
        "window",
        column + 1,
        {str(key): value for key, value in frame.loc[center - pd.Timedelta(hours=3): center + pd.Timedelta(hours=3), column].items()},
    )
for column in range(54, 58):
    series = frame[column].dropna()
    print(
        {
            "column": column + 1,
            "count": len(series),
            "start": str(series.index.min()),
            "end": str(series.index.max()),
            "counts_by_year": series.groupby(series.index.year).count().to_dict(),
            "min": float(series.min()),
            "q01": float(series.quantile(0.01)),
            "median": float(series.median()),
            "q99": float(series.quantile(0.99)),
            "max": float(series.max()),
        }
    )
    print("first", {str(key): value for key, value in series.head().items()})
    print("last", {str(key): value for key, value in series.tail().items()})
